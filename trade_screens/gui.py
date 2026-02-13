#!/usr/bin/env python3
"""PySide6 dark-mode GUI for NSE Stock Investment Analyzer."""

from __future__ import annotations

import os
import re
import sys
import queue
import threading
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
from openpyxl import load_workbook
from PySide6.QtCore import Qt, QTimer, QDateTime
from PySide6.QtGui import QPixmap, QPainter, QColor, QFont
from PySide6.QtCharts import (
    QChart,
    QChartView,
    QCandlestickSeries,
    QCandlestickSet,
    QDateTimeAxis,
    QValueAxis,
    QBarSeries,
    QBarSet,
    QBarCategoryAxis,
    QLineSeries,
)
from PySide6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QComboBox,
    QMainWindow,
    QMessageBox,
    QDialog,
    QPushButton,
    QPlainTextEdit,
    QRadioButton,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

APP_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = APP_DIR.parent
DEFAULT_SYMBOL_FILE = PROJECT_ROOT / "data" / "nse_equity_symbols_eq.txt"
DEFAULT_SYMBOL_CSV = PROJECT_ROOT / "data" / "nse_equity_symbols_eq.csv"


def parse_ticker_tokens(raw: str) -> list[str]:
    tokens = re.split(r"[\s,]+", raw.strip())
    cleaned: list[str] = []
    for token in tokens:
        t = token.strip().upper().replace(".NS", "")
        if t:
            cleaned.append(t)
    return list(dict.fromkeys(cleaned))


def compact_currency(value: float | None) -> str:
    if value is None:
        return "N/A"
    n = float(value)
    a = abs(n)
    if a >= 1_00_00_00_000:
        return f"INR {n/1_00_00_00_000:.2f} Bn"
    if a >= 1_00_00_000:
        return f"INR {n/1_00_00_000:.2f} Cr"
    if a >= 1_00_000:
        return f"INR {n/1_00_000:.2f} L"
    return f"INR {n:,.2f}"


def camel_to_title(name: str) -> str:
    s = re.sub(r"(?<!^)(?=[A-Z])", " ", name)
    return s.strip().title()


def sparkline(values: list[float]) -> str:
    if not values:
        return "N/A"
    blocks = "▁▂▃▄▅▆▇█"
    lo = min(values)
    hi = max(values)
    if hi == lo:
        return blocks[0] * len(values)
    out = []
    for v in values:
        idx = int((v - lo) / (hi - lo) * (len(blocks) - 1))
        out.append(blocks[idx])
    return "".join(out)


def fmt_num(v, digits: int = 2) -> str:
    if v is None:
        return "N/A"
    if isinstance(v, (int, float)):
        return f"{v:,.{digits}f}"
    return str(v)


def fmt_pct(v) -> str:
    if isinstance(v, (int, float)):
        return f"{v * 100:.2f}%"
    return "N/A"


def _get_label_row(ws, label_text: str, max_row: int = 120) -> int | None:
    target = label_text.lower()
    for r in range(1, max_row + 1):
        cell = ws.cell(r, 1).value
        if isinstance(cell, str) and target in cell.lower():
            return r
    return None


def extract_summary_from_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    out: dict = {
        "overview": {},
        "dcf": {},
        "factor": {},
        "risk": {},
        "esg": {},
        "technical": {},
    }

    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        out["overview"] = {
            "Ticker": ws.cell(6, 2).value,
            "Company": ws.cell(6, 3).value,
            "Sector": ws.cell(6, 4).value,
            "Current Price": ws.cell(6, 5).value,
            "Market Cap (Cr)": ws.cell(6, 6).value,
            "P/E": ws.cell(6, 7).value,
            "ROE": ws.cell(6, 8).value,
            "Revenue Growth": ws.cell(6, 9).value,
            "Momentum Score": ws.cell(6, 10).value,
        }

    ticker = out.get("overview", {}).get("Ticker")

    if "DCF Valuation" in wb.sheetnames:
        ws = wb["DCF Valuation"]
        dcf = {}
        target_row = None
        if ticker:
            for r in range(1, 250):
                text = ws.cell(r, 1).value
                if isinstance(text, str) and f"({ticker})" in text:
                    target_row = r
                    break
        if target_row is None:
            target_row = 1
        for r in range(target_row, min(target_row + 60, ws.max_row + 1)):
            label = ws.cell(r, 1).value
            if not isinstance(label, str):
                continue
            l = label.lower()
            if "intrinsic value" in l:
                dcf["Intrinsic Base"] = ws.cell(r, 2).value
                dcf["Intrinsic Bull"] = ws.cell(r, 3).value
                dcf["Intrinsic Bear"] = ws.cell(r, 4).value
            elif "current price" in l:
                dcf["Current Price"] = ws.cell(r, 2).value
            elif "upside" in l:
                dcf["Upside Base"] = ws.cell(r, 2).value
                dcf["Upside Bull"] = ws.cell(r, 3).value
                dcf["Upside Bear"] = ws.cell(r, 4).value
            elif "mean value" in l:
                dcf["MC Mean"] = ws.cell(r, 2).value
            elif "median value" in l:
                dcf["MC Median"] = ws.cell(r, 2).value
            elif "p(upside" in l:
                dcf["MC P(Upside)"] = ws.cell(r, 2).value
        out["dcf"] = dcf

    if "Factor Model" in wb.sheetnames:
        ws = wb["Factor Model"]
        r = 16
        out["factor"] = {
            "Market Beta": ws.cell(r, 2).value,
            "Size Beta": ws.cell(r, 3).value,
            "Value Beta": ws.cell(r, 4).value,
            "Profit Beta": ws.cell(r, 5).value,
            "Invest Beta": ws.cell(r, 6).value,
            "Expected Return": ws.cell(r, 7).value,
            "Excess Return": ws.cell(r, 8).value,
        }

    if "GARCH Risk" in wb.sheetnames:
        ws = wb["GARCH Risk"]
        r = 6
        out["risk"] = {
            "Current Volatility": ws.cell(r, 7).value,
            "Kelly %": ws.cell(r, 8).value,
            "Half Kelly": ws.cell(r, 9).value,
            "VaR 95%": ws.cell(r, 10).value,
            "VaR 99%": ws.cell(r, 11).value,
            "Max Position (L)": ws.cell(r, 12).value,
        }

    if "ESG Scorer" in wb.sheetnames:
        ws = wb["ESG Scorer"]
        r = 12
        out["esg"] = {
            "E Score": ws.cell(r, 2).value,
            "S Score": ws.cell(r, 3).value,
            "G Score": ws.cell(r, 4).value,
            "Weighted ESG": ws.cell(r, 5).value,
            "SDG Align": ws.cell(r, 6).value,
            "Financial Score": ws.cell(r, 7).value,
            "Composite": ws.cell(r, 8).value,
            "Recommendation": ws.cell(r, 9).value,
        }

    if "Technicals" in wb.sheetnames:
        ws = wb["Technicals"]
        r = 4
        out["technical"] = {
            "Current Price": ws.cell(r, 2).value,
            "SMA 50": ws.cell(r, 3).value,
            "SMA 200": ws.cell(r, 4).value,
            "RSI(14)": ws.cell(r, 5).value,
            "MACD": ws.cell(r, 6).value,
            "52W High": ws.cell(r, 7).value,
            "52W Low": ws.cell(r, 8).value,
            "Trend": ws.cell(r, 11).value,
            "RSI Signal": ws.cell(r, 12).value,
            "Momentum": ws.cell(r, 13).value,
        }

    return out


class SummaryDialog(QDialog):
    def __init__(self, summary: dict, report_path: Path, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Analysis Summary")
        self.resize(980, 680)

        layout = QVBoxLayout(self)
        title = QLabel("Model-Wise Analysis Summary")
        title.setObjectName("CardTitle")
        subtitle = QLabel(f"Source report: {report_path.name}")
        subtitle.setStyleSheet("color:#93c5fd;")
        layout.addWidget(title)
        layout.addWidget(subtitle)

        tabs = QTabWidget()
        tabs.addTab(self._build_graph_tab(summary.get("overview", {}), percent_keys={"ROE", "Revenue Growth"}), "Overview")
        tabs.addTab(self._build_graph_tab(summary.get("dcf", {}), percent_keys={"Upside Base", "Upside Bull", "Upside Bear", "MC P(Upside)"}), "DCF")
        tabs.addTab(self._build_graph_tab(summary.get("factor", {}), percent_keys={"Expected Return", "Excess Return"}), "Factor")
        tabs.addTab(self._build_graph_tab(summary.get("risk", {}), percent_keys={"Current Volatility", "Kelly %", "Half Kelly"}), "Risk")
        tabs.addTab(self._build_graph_tab(summary.get("esg", {})), "ESG")
        tabs.addTab(self._build_graph_tab(summary.get("technical", {})), "Technical")
        layout.addWidget(tabs, 1)

        self.setStyleSheet(
            """
            QDialog { background: #0f172a; color: #e2e8f0; }
            QLabel#CardTitle { font-size:18px; font-weight:700; color:#f8fafc; }
            QTabWidget::pane { border:1px solid #334155; background:#111827; border-radius:8px; }
            QTabBar::tab { background:#1f2937; color:#cbd5e1; padding:8px 14px; margin-right:2px; border-top-left-radius:6px; border-top-right-radius:6px; }
            QTabBar::tab:selected { background:#1d4ed8; color:white; }
            QTableWidget { background:#0b1220; color:#e2e8f0; gridline-color:#334155; border:1px solid #334155; }
            QHeaderView::section { background:#1f2937; color:#93c5fd; border:1px solid #334155; padding:6px; }
            """
        )

    def _build_graph_tab(self, data: dict, percent_keys: set[str] | None = None) -> QWidget:
        percent_keys = percent_keys or set()
        w = QWidget()
        l = QVBoxLayout(w)
        numeric_labels: list[str] = []
        numeric_vals: list[float] = []
        detail_rows: list[tuple[str, str]] = []

        rows = list(data.items())
        for k, v in rows:
            if k in percent_keys:
                val_str = fmt_pct(v)
                if isinstance(v, (int, float)):
                    numeric_labels.append(str(k))
                    numeric_vals.append(float(v) * 100.0)
            elif isinstance(v, (int, float)):
                val_str = fmt_num(v)
                numeric_labels.append(str(k))
                numeric_vals.append(float(v))
            else:
                val_str = "N/A" if v is None else str(v)
            detail_rows.append((str(k), val_str))

        chart = QChart()
        chart.setBackgroundVisible(True)
        chart.setBackgroundBrush(QColor("#0b1220"))
        chart.setPlotAreaBackgroundVisible(True)
        chart.setPlotAreaBackgroundBrush(QColor("#0b1220"))
        chart.legend().hide()

        if numeric_vals:
            bar_set = QBarSet("Value")
            bar_set.setColor(QColor("#3b82f6"))
            for val in numeric_vals:
                bar_set.append(val)
            series = QBarSeries()
            series.append(bar_set)
            series.setLabelsVisible(True)
            chart.addSeries(series)

            axis_x = QBarCategoryAxis()
            axis_x.append(numeric_labels)
            axis_x.setLabelsColor(QColor("#cbd5e1"))
            axis_y = QValueAxis()
            ymin = min(0.0, min(numeric_vals) * 1.15)
            ymax = max(numeric_vals) * 1.15 if max(numeric_vals) != 0 else 1.0
            axis_y.setRange(ymin, ymax)
            axis_y.setLabelsColor(QColor("#cbd5e1"))
            chart.addAxis(axis_x, Qt.AlignBottom)
            chart.addAxis(axis_y, Qt.AlignLeft)
            series.attachAxis(axis_x)
            series.attachAxis(axis_y)
            chart.setTitle("Bar Chart Summary")
            chart.setTitleBrush(QColor("#93c5fd"))
        else:
            chart.setTitle("No numeric metrics available for chart rendering")
            chart.setTitleBrush(QColor("#93c5fd"))

        chart_view = QChartView(chart)
        chart_view.setMinimumHeight(300)
        l.addWidget(chart_view)

        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Metric", "Value"])
        table.setRowCount(len(detail_rows))
        for i, (k, val) in enumerate(detail_rows):
            table.setItem(i, 0, QTableWidgetItem(k))
            table.setItem(i, 1, QTableWidgetItem(val))
        table.horizontalHeader().setStretchLastSection(True)
        table.verticalHeader().setVisible(False)
        l.addWidget(table)
        return w


class AnalyzerWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Trade_Screens terminal")
        self.resize(1360, 860)

        self.mode = "single"
        self.proc: subprocess.Popen | None = None
        self.event_queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.profile_job_id = 0
        self.chart_job_id = 0
        self.current_symbol: Optional[str] = None
        self.last_report_path: Optional[Path] = None
        self.run_started_at: Optional[datetime] = None
        self.symbol_rows = self._load_symbol_rows()

        self._build_ui()
        self._apply_dark_theme()
        self._refresh_symbol_list()
        self._update_terminal_clock()
        # Make summary available if a recent report already exists.
        existing = self._find_latest_report()
        if existing and existing.exists():
            self.last_report_path = existing
            self.summary_btn.setEnabled(True)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self._drain_events)
        self.timer.start(120)
        self.clock_timer = QTimer(self)
        self.clock_timer.timeout.connect(self._update_terminal_clock)
        self.clock_timer.start(1000)

    def _build_ui(self) -> None:
        root = QWidget()
        self.setCentralWidget(root)
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(8, 8, 8, 8)
        root_layout.setSpacing(6)

        terminal_strip = QFrame()
        terminal_strip.setObjectName("TerminalStrip")
        strip_layout = QHBoxLayout(terminal_strip)
        strip_layout.setContentsMargins(10, 4, 10, 4)
        strip_layout.setSpacing(14)
        self.market_mode_label = QLabel("MODE: NSE EQUITY")
        self.market_mode_label.setObjectName("TerminalStripLabel")
        self.market_feed_label = QLabel("FEED: YAHOO FINANCE")
        self.market_feed_label.setObjectName("TerminalStripLabel")
        self.market_clock_label = QLabel("LOCAL 00 Jan 0000 00:00:00")
        self.market_clock_label.setObjectName("TerminalStripLabel")
        strip_layout.addWidget(self.market_mode_label)
        strip_layout.addWidget(self.market_feed_label)
        strip_layout.addStretch(1)
        strip_layout.addWidget(self.market_clock_label)
        root_layout.addWidget(terminal_strip)

        header = QLabel("TRADE_SCREENS EQUITY TERMINAL")
        header.setObjectName("Header")
        sub = QLabel("Bloomberg-style dashboard | Press Esc to exit full-screen")
        sub.setObjectName("SubHeader")
        root_layout.addWidget(header)
        root_layout.addWidget(sub)

        mode_card = QFrame()
        mode_card.setObjectName("Card")
        mode_layout = QVBoxLayout(mode_card)

        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Search Approach:"))
        self.rb_single = QRadioButton("Selected Stock")
        self.rb_multi = QRadioButton("Multiple Stocks")
        self.rb_file = QRadioButton("From File")
        self.rb_single.setChecked(True)
        self.mode_group = QButtonGroup(self)
        for rb in (self.rb_single, self.rb_multi, self.rb_file):
            self.mode_group.addButton(rb)
            mode_row.addWidget(rb)
        mode_row.addStretch(1)
        mode_layout.addLayout(mode_row)

        self.single_row = QHBoxLayout()
        self.single_row.addWidget(QLabel("Ticker:"))
        self.single_input = QLineEdit()
        self.single_input.setPlaceholderText("Example: KAYNES")
        self.single_row.addWidget(self.single_input)
        mode_layout.addLayout(self.single_row)

        self.multi_row = QVBoxLayout()
        self.multi_row.addWidget(QLabel("Tickers (comma/space/newline separated):"))
        self.multi_input = QPlainTextEdit()
        self.multi_input.setFixedHeight(80)
        self.multi_row.addWidget(self.multi_input)
        mode_layout.addLayout(self.multi_row)
        self._set_layout_visible(self.multi_row, False)

        self.file_row = QHBoxLayout()
        self.file_row.addWidget(QLabel("Ticker File:"))
        self.file_input = QLineEdit(str(DEFAULT_SYMBOL_FILE))
        self.file_row.addWidget(self.file_input)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self._browse_file)
        self.file_row.addWidget(browse_btn)
        mode_layout.addLayout(self.file_row)
        self._set_layout_visible(self.file_row, False)

        self.rb_single.toggled.connect(lambda checked: checked and self._switch_mode("single"))
        self.rb_multi.toggled.connect(lambda checked: checked and self._switch_mode("multi"))
        self.rb_file.toggled.connect(lambda checked: checked and self._switch_mode("file"))

        root_layout.addWidget(mode_card)

        splitter = QSplitter(Qt.Horizontal)
        root_layout.addWidget(splitter, 1)

        left_panel = QFrame()
        left_panel.setObjectName("Card")
        left_layout = QVBoxLayout(left_panel)
        left_layout.addWidget(QLabel("NSE Stock Universe"))
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Filter by symbol or company name")
        self.filter_input.textChanged.connect(self._refresh_symbol_list)
        left_layout.addWidget(self.filter_input)

        self.stock_list = QListWidget()
        self.stock_list.setSelectionMode(QListWidget.ExtendedSelection)
        self.stock_list.itemSelectionChanged.connect(self._on_stock_selection_changed)
        left_layout.addWidget(self.stock_list, 1)

        stock_btns = QHBoxLayout()
        btn_single = QPushButton("Use for Single")
        btn_multi = QPushButton("Add to Multiple")
        btn_single.clicked.connect(self._use_for_single)
        btn_multi.clicked.connect(self._add_to_multi)
        stock_btns.addWidget(btn_single)
        stock_btns.addWidget(btn_multi)
        left_layout.addLayout(stock_btns)

        splitter.addWidget(left_panel)

        right_split = QSplitter(Qt.Vertical)
        splitter.addWidget(right_split)
        splitter.setSizes([520, 1320])

        profile_card = QFrame()
        profile_card.setObjectName("Card")
        profile_layout = QVBoxLayout(profile_card)
        profile_title = QLabel("Selected Company Profile")
        profile_title.setObjectName("CardTitle")
        profile_layout.addWidget(profile_title)

        top_row = QHBoxLayout()
        self.logo_label = QLabel()
        self.logo_label.setFixedSize(84, 84)
        self.logo_label.setPixmap(self._build_placeholder_logo("NSE"))
        self.logo_label.setAlignment(Qt.AlignCenter)
        top_row.addWidget(self.logo_label, 0, Qt.AlignTop)

        text_col = QVBoxLayout()
        self.profile_symbol = QLabel("Symbol: -")
        self.profile_company = QLabel("Company: -")
        self.profile_sector = QLabel("Sector: -")
        self.profile_website = QLabel("Website: -")
        self.profile_website.setOpenExternalLinks(True)
        text_col.addWidget(self.profile_symbol)
        text_col.addWidget(self.profile_company)
        text_col.addWidget(self.profile_sector)
        text_col.addWidget(self.profile_website)
        top_row.addLayout(text_col, 1)
        profile_layout.addLayout(top_row)

        metrics_grid = QGridLayout()
        self.profile_price = QLabel("Current Price: -")
        self.profile_mcap = QLabel("Market Cap: -")
        self.profile_pe = QLabel("P/E: -")
        self.profile_earnings_date = QLabel("Nearest Earnings Date: -")
        self.profile_qoq = QLabel("QoQ Earnings Growth: -")
        self.profile_yoy = QLabel("YoY Earnings Growth: -")
        self.profile_status = QLabel("Status: Select a stock to load preview")
        metrics_grid.addWidget(self.profile_price, 0, 0)
        metrics_grid.addWidget(self.profile_mcap, 0, 1)
        metrics_grid.addWidget(self.profile_pe, 1, 0)
        metrics_grid.addWidget(self.profile_earnings_date, 1, 1)
        metrics_grid.addWidget(self.profile_qoq, 2, 0)
        metrics_grid.addWidget(self.profile_yoy, 2, 1)
        metrics_grid.addWidget(self.profile_status, 3, 0, 1, 2)
        profile_layout.addLayout(metrics_grid)

        self.profile_history = QPlainTextEdit()
        self.profile_history.setReadOnly(True)
        self.profile_history.setFixedHeight(85)
        self.profile_history.setPlaceholderText("Brief company history appears here.")
        profile_layout.addWidget(self.profile_history)

        quarterly_title = QLabel("Quarterly Results (Last 1 Year)")
        quarterly_title.setObjectName("CardTitle")
        profile_layout.addWidget(quarterly_title)
        self.quarterly_table = QTableWidget()
        self.quarterly_table.setMinimumHeight(220)
        self.quarterly_table.setColumnCount(1)
        self.quarterly_table.setHorizontalHeaderLabels(["Metric"])
        self.quarterly_table.verticalHeader().setVisible(False)
        profile_layout.addWidget(self.quarterly_table)

        right_split.addWidget(profile_card)

        chart_card = QFrame()
        chart_card.setObjectName("Card")
        chart_layout = QVBoxLayout(chart_card)
        chart_title = QLabel("Price Movement (Candlestick)")
        chart_title.setObjectName("CardTitle")
        chart_layout.addWidget(chart_title)

        chart_ctrl = QHBoxLayout()
        chart_ctrl.addWidget(QLabel("Range"))
        self.range_combo = QComboBox()
        self.range_combo.addItems(["1M", "3M", "6M", "1Y"])
        self.range_combo.setCurrentText("1M")
        chart_ctrl.addWidget(self.range_combo)
        chart_ctrl.addWidget(QLabel("Interval"))
        self.interval_combo = QComboBox()
        self.interval_combo.addItems(["1d", "1h", "30m"])
        self.interval_combo.setCurrentText("1d")
        chart_ctrl.addWidget(self.interval_combo)
        self.reload_chart_btn = QPushButton("Reload Chart")
        chart_ctrl.addWidget(self.reload_chart_btn)
        chart_ctrl.addStretch(1)
        self.chart_key_label = QLabel("Keys: Green=Bullish Candle | Red=Bearish Candle")
        chart_ctrl.addWidget(self.chart_key_label)
        chart_layout.addLayout(chart_ctrl)

        self.candle_chart = QChart()
        self.candle_chart.setBackgroundVisible(True)
        self.candle_chart.setBackgroundBrush(QColor("#0b1220"))
        self.candle_chart.setPlotAreaBackgroundVisible(True)
        self.candle_chart.setPlotAreaBackgroundBrush(QColor("#0b1220"))
        self.candle_chart.legend().hide()
        self.candle_view = QChartView(self.candle_chart)
        self.candle_view.setMinimumHeight(280)
        chart_layout.addWidget(self.candle_view)
        right_split.addWidget(chart_card)

        chat_card = QFrame()
        chat_card.setObjectName("Card")
        chat_layout = QVBoxLayout(chat_card)
        chat_title = QLabel("Analysis Chat View")
        chat_title.setObjectName("CardTitle")
        chat_layout.addWidget(chat_title)

        self.chat_box = QPlainTextEdit()
        self.chat_box.setReadOnly(True)
        chat_layout.addWidget(self.chat_box, 1)
        right_split.addWidget(chat_card)
        right_split.setSizes([390, 390, 250])

        self.range_combo.currentTextChanged.connect(lambda _v: self._reload_chart_for_current())
        self.interval_combo.currentTextChanged.connect(lambda _v: self._reload_chart_for_current())
        self.reload_chart_btn.clicked.connect(self._reload_chart_for_current)

        controls = QHBoxLayout()
        self.run_btn = QPushButton("RUN ANALYSIS")
        self.run_btn.clicked.connect(self._run_analysis)
        self.summary_btn = QPushButton("VIEW ANALYSIS SUMMARY")
        self.summary_btn.setEnabled(False)
        self.summary_btn.clicked.connect(self._show_summary_dialog)
        clear_btn = QPushButton("CLEAR FEED")
        clear_btn.clicked.connect(lambda: self.chat_box.setPlainText(""))
        controls.addWidget(self.run_btn)
        controls.addWidget(self.summary_btn)
        controls.addWidget(clear_btn)
        controls.addStretch(1)
        self.status_label = QLabel("IDLE")
        self.status_label.setObjectName("StatusLabel")
        controls.addWidget(self.status_label)
        root_layout.addLayout(controls)

    def _apply_dark_theme(self) -> None:
        self.setStyleSheet(
            """
            QWidget {
                background: #0b0d10;
                color: #f4f5f7;
                font-family: Consolas;
                font-size: 12px;
            }
            QFrame#TerminalStrip {
                background: #121418;
                border: 1px solid #f0b90b;
            }
            QLabel#TerminalStripLabel {
                color: #f0b90b;
                font-size: 11px;
                font-weight: 600;
            }
            QFrame#Card {
                background: #121418;
                border: 1px solid #2b3138;
                border-radius: 0px;
                padding: 6px;
            }
            QLabel#Header {
                font-size: 22px;
                font-weight: 700;
                color: #ffb000;
                letter-spacing: 1px;
            }
            QLabel#SubHeader {
                color: #9aa3ad;
                font-size: 11px;
            }
            QLabel#CardTitle {
                font-size: 13px;
                font-weight: 600;
                color: #ffbf47;
                text-transform: uppercase;
            }
            QLabel#StatusLabel {
                color: #ffbf47;
                font-weight: 700;
            }
            QLineEdit, QPlainTextEdit, QListWidget, QComboBox, QTableWidget {
                background: #090b0d;
                border: 1px solid #2f343d;
                border-radius: 0px;
                padding: 5px;
                selection-background-color: #174a9c;
                selection-color: #ffffff;
                alternate-background-color: #12151a;
            }
            QComboBox::drop-down {
                border: 0px;
                width: 18px;
            }
            QPushButton {
                background: #1a1d23;
                border: 1px solid #ffbf47;
                border-radius: 0px;
                padding: 7px 10px;
                color: #ffbf47;
                font-weight: 600;
                text-transform: uppercase;
            }
            QPushButton:hover { background: #252a32; }
            QPushButton:pressed { background: #313846; }
            QPushButton:disabled { background: #11141a; border-color: #39414d; color: #6f7a87; }
            QTableWidget {
                gridline-color: #2e343d;
            }
            QHeaderView::section {
                background: #101318;
                color: #f0b90b;
                border: 1px solid #2f3640;
                padding: 4px;
                font-weight: 700;
            }
            QListWidget::item {
                padding: 5px 3px;
            }
            QListWidget::item:selected {
                background: #213758;
                color: #ffffff;
            }
            QSplitter::handle {
                background: #242a31;
            }
            QRadioButton::indicator {
                width: 14px;
                height: 14px;
            }
            """
        )

        self.quarterly_table.setAlternatingRowColors(True)
        self.stock_list.setAlternatingRowColors(True)

    def _update_terminal_clock(self) -> None:
        now = QDateTime.currentDateTime().toString("dd MMM yyyy HH:mm:ss")
        self.market_clock_label.setText(f"LOCAL {now}")

    def keyPressEvent(self, event) -> None:
        if event.key() == Qt.Key_Escape and self.isFullScreen():
            self.showMaximized()
            return
        super().keyPressEvent(event)

    def _load_symbol_rows(self) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        if DEFAULT_SYMBOL_CSV.exists():
            try:
                df = pd.read_csv(DEFAULT_SYMBOL_CSV)
                symbol_col = next((c for c in df.columns if c.strip().upper() == "SYMBOL"), None)
                name_col = next((c for c in df.columns if c.strip().upper() == "NAME OF COMPANY"), None)
                if symbol_col:
                    for _, row in df.iterrows():
                        sym = str(row.get(symbol_col, "")).strip().upper()
                        if not sym:
                            continue
                        name = str(row.get(name_col, "")).strip() if name_col else ""
                        rows.append({"symbol": sym, "name": name})
            except Exception:
                rows = []

        if rows:
            uniq = {}
            for r in rows:
                uniq[r["symbol"]] = r
            return [uniq[k] for k in sorted(uniq.keys())]

        if DEFAULT_SYMBOL_FILE.exists():
            syms = []
            for line in DEFAULT_SYMBOL_FILE.read_text(encoding="utf-8").splitlines():
                sym = line.strip().upper().replace(".NS", "")
                if sym:
                    syms.append(sym)
            return [{"symbol": s, "name": ""} for s in sorted(set(syms))]

        return []

    def _refresh_symbol_list(self) -> None:
        q = self.filter_input.text().strip().upper()
        self.stock_list.clear()
        for row in self.symbol_rows:
            sym = row["symbol"]
            name = row["name"]
            combined = f"{sym} | {name}" if name else sym
            if q and q not in combined.upper():
                continue
            item = QListWidgetItem(combined)
            item.setData(Qt.UserRole, sym)
            item.setToolTip(name or sym)
            self.stock_list.addItem(item)

    def _selected_symbols(self) -> list[str]:
        selected = self.stock_list.selectedItems()
        return [item.data(Qt.UserRole) for item in selected]

    def _use_for_single(self) -> None:
        picks = self._selected_symbols()
        if not picks:
            return
        sym = picks[0]
        self.current_symbol = sym
        self.rb_single.setChecked(True)
        self.single_input.setText(sym)
        self._append_chat(f"Selected stock: {sym}")
        self._start_profile_fetch(sym)
        self._start_candles_fetch(sym)

    def _add_to_multi(self) -> None:
        picks = self._selected_symbols()
        if not picks:
            return
        self.current_symbol = picks[0]
        self.rb_multi.setChecked(True)
        existing = parse_ticker_tokens(self.multi_input.toPlainText())
        merged = list(dict.fromkeys(existing + picks))
        self.multi_input.setPlainText(", ".join(merged))
        self._append_chat(f"Added stocks: {', '.join(picks)}")
        self._start_profile_fetch(picks[0])
        self._start_candles_fetch(picks[0])

    def _switch_mode(self, mode: str) -> None:
        self.mode = mode
        self._set_layout_visible(self.single_row, mode == "single")
        self._set_layout_visible(self.multi_row, mode == "multi")
        self._set_layout_visible(self.file_row, mode == "file")

    def _set_layout_visible(self, layout, visible: bool) -> None:
        for i in range(layout.count()):
            item = layout.itemAt(i)
            widget = item.widget()
            child_layout = item.layout()
            if widget:
                widget.setVisible(visible)
            elif child_layout:
                self._set_layout_visible(child_layout, visible)

    def _browse_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select ticker file",
            str(PROJECT_ROOT),
            "Ticker Files (*.txt *.csv);;All Files (*.*)",
        )
        if path:
            self.file_input.setText(path)
            self._append_chat(f"Selected ticker file: {path}")

    def _on_stock_selection_changed(self) -> None:
        picks = self._selected_symbols()
        if picks:
            self.current_symbol = picks[0]
            self._start_profile_fetch(picks[0])
            self._start_candles_fetch(picks[0])

    def _reload_chart_for_current(self) -> None:
        if self.current_symbol:
            self._start_candles_fetch(self.current_symbol)

    def _range_to_query(self, label: str) -> str:
        return {
            "1M": "1mo",
            "3M": "3mo",
            "6M": "6mo",
            "1Y": "1y",
        }.get(label, "1mo")

    def _find_latest_report(self, since: Optional[datetime] = None) -> Optional[Path]:
        report_dir = PROJECT_ROOT / "reports"
        files = []
        if report_dir.exists():
            files.extend(report_dir.glob("NSE_Analysis_*.xlsx"))
        files.extend(PROJECT_ROOT.glob("NSE_Analysis_*.xlsx"))
        files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
        if not files:
            return None
        if since is None:
            return files[0]
        for p in files:
            ts = datetime.fromtimestamp(p.stat().st_mtime)
            if ts >= since:
                return p
        return files[0]

    def _start_candles_fetch(self, symbol: str) -> None:
        self.chart_job_id += 1
        job_id = self.chart_job_id
        self.chart_key_label.setText("Keys: Loading chart...")
        rng = self._range_to_query(self.range_combo.currentText())
        interval = self.interval_combo.currentText()

        def worker() -> None:
            candles = self._fetch_candles(symbol, rng, interval)
            self.event_queue.put(("candles", (job_id, symbol, rng, interval, candles)))

        threading.Thread(target=worker, daemon=True).start()

    def _fetch_candles(self, symbol: str, rng: str, interval: str) -> list[dict]:
        s = requests.Session()
        s.trust_env = False
        s.headers.update({"User-Agent": "Mozilla/5.0", "Cache-Control": "no-cache"})
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}.NS"
        def do_fetch(req_interval: str) -> list[dict]:
            resp = s.get(url, params={"range": rng, "interval": req_interval}, timeout=20)
            if resp.status_code != 200:
                return []
            result = (resp.json().get("chart", {}).get("result") or [{}])[0]
            timestamps = result.get("timestamp") or []
            quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
            opens = quote.get("open") or []
            highs = quote.get("high") or []
            lows = quote.get("low") or []
            closes = quote.get("close") or []
            volumes = quote.get("volume") or []
            candles = []
            for t, o, h, l, c, v in zip(timestamps, opens, highs, lows, closes, volumes):
                if not all(isinstance(v, (int, float)) for v in (o, h, l, c)):
                    continue
                candles.append({
                    "ts": int(t),
                    "open": float(o),
                    "high": float(h),
                    "low": float(l),
                    "close": float(c),
                    "volume": float(v) if isinstance(v, (int, float)) else 0.0,
                })
            return candles

        try:
            primary = do_fetch(interval)
            if primary:
                return primary
            if interval != "1d":
                return do_fetch("1d")
            return []
        except Exception:
            return []

    def _render_candles(self, symbol: str, rng: str, interval: str, candles: list[dict]) -> None:
        self.candle_chart.removeAllSeries()
        for ax in self.candle_chart.axes():
            self.candle_chart.removeAxis(ax)

        if not candles:
            self.chart_key_label.setText(
                f"Keys: Green=Bullish Candle | Red=Bearish Candle | {symbol} {rng}/{interval}: no data"
            )
            return

        series = QCandlestickSeries()
        series.setIncreasingColor(QColor("#16a34a"))
        series.setDecreasingColor(QColor("#dc2626"))
        series.setBodyWidth(0.7)
        vol_series = QLineSeries()
        vol_series.setName("Volume")
        vol_series.setColor(QColor("#f59e0b"))

        min_price = float("inf")
        max_price = float("-inf")
        min_ts = candles[0]["ts"]
        max_ts = candles[-1]["ts"]
        max_vol = 0.0
        for row in candles:
            s = QCandlestickSet(row["open"], row["high"], row["low"], row["close"], row["ts"] * 1000)
            series.append(s)
            vol_series.append(row["ts"] * 1000, row.get("volume", 0.0))
            min_price = min(min_price, row["low"])
            max_price = max(max_price, row["high"])
            max_vol = max(max_vol, row.get("volume", 0.0))

        self.candle_chart.addSeries(series)
        self.candle_chart.addSeries(vol_series)
        x_axis = QDateTimeAxis()
        x_axis.setFormat("dd MMM")
        x_axis.setTitleText("Time")
        x_axis.setLabelsColor(QColor("#cbd5e1"))
        x_axis.setTitleBrush(QColor("#cbd5e1"))
        x_axis.setRange(
            QDateTime.fromSecsSinceEpoch(min_ts),
            QDateTime.fromSecsSinceEpoch(max_ts),
        )
        y_axis = QValueAxis()
        y_axis.setTitleText("Price (INR)")
        y_axis.setLabelsColor(QColor("#cbd5e1"))
        y_axis.setTitleBrush(QColor("#cbd5e1"))
        y_axis_vol = QValueAxis()
        y_axis_vol.setTitleText("Volume")
        y_axis_vol.setLabelsColor(QColor("#f59e0b"))
        y_axis_vol.setTitleBrush(QColor("#f59e0b"))
        pad = (max_price - min_price) * 0.08 if max_price > min_price else max_price * 0.02
        y_axis.setRange(min_price - pad, max_price + pad)
        y_axis_vol.setRange(0.0, max_vol * 1.15 if max_vol > 0 else 1.0)

        self.candle_chart.addAxis(x_axis, Qt.AlignBottom)
        self.candle_chart.addAxis(y_axis, Qt.AlignLeft)
        self.candle_chart.addAxis(y_axis_vol, Qt.AlignRight)
        series.attachAxis(x_axis)
        series.attachAxis(y_axis)
        vol_series.attachAxis(x_axis)
        vol_series.attachAxis(y_axis_vol)

        first_close = candles[0]["close"]
        last_close = candles[-1]["close"]
        change = ((last_close / first_close) - 1) * 100 if first_close else 0
        last_vol = candles[-1].get("volume", 0.0)
        self.chart_key_label.setText(
            f"Keys: Green=Bullish | Red=Bearish | O:{candles[-1]['open']:.2f} H:{candles[-1]['high']:.2f} "
            f"L:{candles[-1]['low']:.2f} C:{candles[-1]['close']:.2f} | Vol:{last_vol:,.0f} | Δ {change:+.2f}%"
        )

    def _start_profile_fetch(self, symbol: str) -> None:
        self.profile_job_id += 1
        job_id = self.profile_job_id
        self.profile_status.setText("Status: Loading profile...")

        def worker() -> None:
            profile = self._fetch_profile(symbol)
            self.event_queue.put(("profile", (job_id, profile)))

        threading.Thread(target=worker, daemon=True).start()

    def _lookup_symbol_name(self, symbol: str) -> str:
        for row in self.symbol_rows:
            if row["symbol"] == symbol:
                return row.get("name", "") or ""
        return ""

    def _fetch_profile(self, symbol: str) -> dict:
        s = requests.Session()
        s.trust_env = False
        s.headers.update({"User-Agent": "Mozilla/5.0"})

        fallback_name = self._lookup_symbol_name(symbol)
        profile = {
            "symbol": symbol,
            "company": fallback_name or "N/A",
            "sector": "N/A",
            "price": None,
            "market_cap": None,
            "pe": None,
            "website": None,
            "history": None,
            "chart_points": [],
            "chart_dates": [],
            "qoq_growth": None,
            "yoy_growth": None,
            "nearest_earnings_date": None,
            "quarter_labels": [],
            "quarterly_results": {},
            "logo_pixmap": None,
        }

        try:
            resp = s.get(
                "https://query2.finance.yahoo.com/v1/finance/search",
                params={"q": f"{symbol}.NS", "quotesCount": 1, "newsCount": 0},
                timeout=18,
            )
            if resp.status_code == 200:
                quotes = resp.json().get("quotes") or []
                if quotes:
                    q = quotes[0]
                    profile["company"] = q.get("longname") or q.get("shortname") or profile["company"]
                    profile["sector"] = q.get("sectorDisp") or q.get("sector") or profile["sector"]
        except Exception:
            pass

        try:
            resp = s.get(
                f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}.NS",
                params={"range": "1mo", "interval": "1d"},
                timeout=18,
            )
            if resp.status_code == 200:
                result = (resp.json().get("chart", {}).get("result") or [{}])[0]
                meta = result.get("meta") or {}
                profile["price"] = meta.get("regularMarketPrice")
                if profile["company"] == "N/A":
                    profile["company"] = meta.get("longName") or meta.get("shortName") or profile["company"]
                timestamps = result.get("timestamp") or []
                quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
                closes = quote.get("close") or []
                pairs = [(t, c) for t, c in zip(timestamps, closes) if isinstance(c, (int, float))]
                profile["chart_dates"] = [datetime.fromtimestamp(t).strftime("%d-%b") for t, _ in pairs]
                profile["chart_points"] = [float(c) for _, c in pairs]
        except Exception:
            pass

        shares = None
        net_income_ttm = None
        ni_series: list[tuple[Optional[str], float]] = []
        try:
            resp = s.get(
                f"https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/{symbol}.NS",
                params={
                    "type": "quarterlyBasicAverageShares,quarterlyNetIncome",
                    "merge": "false",
                    "period1": "1609459200",
                    "period2": str(int(datetime.now().timestamp())),
                },
                timeout=18,
            )
            if resp.status_code == 200:
                result = resp.json().get("timeseries", {}).get("result") or []
                for item in result:
                    t = ((item.get("meta") or {}).get("type") or [None])[0]
                    if t == "quarterlyBasicAverageShares":
                        vals = [x.get("reportedValue", {}).get("raw") for x in item.get(t, []) if isinstance(x, dict)]
                        vals = [v for v in vals if isinstance(v, (int, float))]
                        if vals:
                            shares = vals[-1]
                    if t == "quarterlyNetIncome":
                        raw_items = [x for x in item.get(t, []) if isinstance(x, dict)]
                        vals = [x.get("reportedValue", {}).get("raw") for x in raw_items]
                        vals = [v for v in vals if isinstance(v, (int, float))]
                        if vals:
                            net_income_ttm = sum(vals[-4:]) if len(vals) >= 4 else vals[-1]
                        for x in raw_items:
                            rv = x.get("reportedValue", {}).get("raw")
                            ad = x.get("asOfDate")
                            if isinstance(rv, (int, float)):
                                ni_series.append((ad, float(rv)))
        except Exception:
            pass

        if profile["price"] and shares:
            profile["market_cap"] = float(profile["price"]) * float(shares)
        if profile["market_cap"] and net_income_ttm and net_income_ttm > 0:
            profile["pe"] = float(profile["market_cap"]) / float(net_income_ttm)

        ni_series.sort(key=lambda x: x[0] or "")
        ni_vals = [v for _, v in ni_series]
        if len(ni_vals) >= 2 and ni_vals[-2] != 0:
            profile["qoq_growth"] = (ni_vals[-1] / ni_vals[-2]) - 1
        if len(ni_vals) >= 5 and ni_vals[-5] != 0:
            profile["yoy_growth"] = (ni_vals[-1] / ni_vals[-5]) - 1
        if ni_series and ni_series[-1][0]:
            try:
                last_dt = datetime.strptime(ni_series[-1][0], "%Y-%m-%d")
                # Estimate next earnings publication roughly 45 days after next quarter end.
                est = last_dt.replace(day=1)
                month = est.month + 3
                year = est.year + (month - 1) // 12
                month = ((month - 1) % 12) + 1
                est = est.replace(year=year, month=month)
                profile["nearest_earnings_date"] = est.strftime("%d %b %Y") + " (estimated)"
            except Exception:
                profile["nearest_earnings_date"] = None

        website, domain = self._fetch_official_website(s, profile["company"], symbol)
        profile["website"] = website
        profile["history"] = self._build_company_history(profile, fallback_name)
        q_labels, q_results = self._fetch_quarterly_results(s, symbol)
        profile["quarter_labels"] = q_labels
        profile["quarterly_results"] = q_results
        profile["logo_pixmap"] = self._fetch_logo_pixmap(s, domain, profile["company"], symbol)
        return profile

    def _fetch_quarterly_results(self, session: requests.Session, symbol: str) -> tuple[list[str], dict[str, dict[str, float]]]:
        # Pull a wide set of quarterly metrics; render whatever is available.
        metric_types = [
            "quarterlyTotalRevenue",
            "quarterlyCostOfRevenue",
            "quarterlyGrossProfit",
            "quarterlyOperatingIncome",
            "quarterlyEBITDA",
            "quarterlyNetIncome",
            "quarterlyDilutedEPS",
            "quarterlyBasicEPS",
            "quarterlyTotalExpenses",
            "quarterlySellingGeneralAndAdministration",
            "quarterlyResearchAndDevelopment",
            "quarterlyTaxProvision",
            "quarterlyPretaxIncome",
            "quarterlyOperatingCashFlow",
            "quarterlyFreeCashFlow",
            "quarterlyCapitalExpenditure",
            "quarterlyTotalDebt",
            "quarterlyCashAndCashEquivalents",
            "quarterlyCurrentAssets",
            "quarterlyCurrentLiabilities",
            "quarterlyStockholdersEquity",
            "quarterlyBasicAverageShares",
        ]
        try:
            resp = session.get(
                f"https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/{symbol}.NS",
                params={
                    "type": ",".join(metric_types),
                    "merge": "false",
                    "period1": "1609459200",
                    "period2": str(int(datetime.now().timestamp())),
                },
                timeout=20,
            )
            if resp.status_code != 200:
                return [], {}
            result = resp.json().get("timeseries", {}).get("result") or []
            metrics_map: dict[str, dict[str, float]] = {}
            all_dates: set[str] = set()
            for item in result:
                typ = ((item.get("meta") or {}).get("type") or [None])[0]
                if not typ:
                    continue
                rows = item.get(typ) or []
                metric_values: dict[str, float] = {}
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    asof = row.get("asOfDate")
                    val = row.get("reportedValue", {}).get("raw")
                    if asof and isinstance(val, (int, float)):
                        metric_values[asof] = float(val)
                        all_dates.add(asof)
                if metric_values:
                    friendly = camel_to_title(typ.replace("quarterly", "", 1))
                    metrics_map[friendly] = metric_values
            if not all_dates:
                return [], {}
            quarter_labels = sorted(all_dates)[-4:]  # last 1 year (4 quarters)
            return quarter_labels, metrics_map
        except Exception:
            return [], {}

    def _fetch_official_website(self, session: requests.Session, company_name: str, symbol: str) -> tuple[Optional[str], Optional[str]]:
        domain = None
        try:
            resp = session.get(
                "https://autocomplete.clearbit.com/v1/companies/suggest",
                params={"query": company_name if company_name and company_name != "N/A" else symbol},
                timeout=12,
            )
            if resp.status_code == 200:
                arr = resp.json() or []
                if arr:
                    domain = arr[0].get("domain")
        except Exception:
            domain = None

        if domain:
            return f"https://{domain}", domain
        return None, None

    def _build_company_history(self, profile: dict, fallback_name: str) -> str:
        company = profile.get("company") or fallback_name or profile.get("symbol") or "This company"
        sector = profile.get("sector") or "Unknown sector"
        price = profile.get("price")
        mcap = profile.get("market_cap")
        pe = profile.get("pe")
        pieces = [
            f"{company} is an NSE-listed company operating in the {sector} sector.",
        ]
        if price is not None:
            pieces.append(f"Current market price is {compact_currency(price)}.")
        if mcap is not None:
            pieces.append(f"Current market capitalization is {compact_currency(mcap)}.")
        if pe is not None:
            pieces.append(f"The stock trades around {pe:.2f}x earnings.")
        pieces.append("This profile summary is generated from currently available market endpoints.")
        return " ".join(pieces)

    def _fetch_logo_pixmap(
        self,
        session: requests.Session,
        domain: Optional[str],
        company_name: str,
        symbol: str
    ) -> QPixmap:
        if domain:
            # Try official site favicon first.
            for logo_url in (
                f"https://{domain}/favicon.ico",
                f"https://www.{domain}/favicon.ico",
                f"https://www.google.com/s2/favicons?domain={domain}&sz=256",
            ):
                try:
                    resp = session.get(logo_url, timeout=12)
                    if resp.status_code == 200 and resp.content:
                        pix = QPixmap()
                        if pix.loadFromData(resp.content):
                            return pix.scaled(84, 84, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                except Exception:
                    pass

        # Last internet attempt based on company query domain guess.
        website, domain2 = self._fetch_official_website(session, company_name, symbol)
        if domain2:
            try:
                fav = session.get(f"https://www.google.com/s2/favicons?domain={domain2}&sz=256", timeout=12)
                if fav.status_code == 200 and fav.content:
                    pix = QPixmap()
                    if pix.loadFromData(fav.content):
                        return pix.scaled(84, 84, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            except Exception:
                pass

        return self._build_placeholder_logo(symbol)

    def _build_placeholder_logo(self, text: str) -> QPixmap:
        initials = (text or "NA")[:3].upper()
        pix = QPixmap(84, 84)
        pix.fill(Qt.transparent)
        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setBrush(QColor("#1d4ed8"))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(0, 0, 84, 84, 16, 16)
        painter.setPen(QColor("#eff6ff"))
        font = QFont("Segoe UI", 20, QFont.Bold)
        painter.setFont(font)
        painter.drawText(pix.rect(), Qt.AlignCenter, initials)
        painter.end()
        return pix

    def _build_cmd(self) -> list[str]:
        cmd = [sys.executable, "-u", "-m", "trade_screens.analyzer"]

        if self.mode == "single":
            ticker = self.single_input.text().strip().upper().replace(".NS", "")
            if not ticker:
                raise ValueError("Provide one ticker for selected stock search.")
            self._append_chat(f"Search mode: selected stock | {ticker}")
            cmd += ["--single", ticker]

        elif self.mode == "multi":
            tickers = parse_ticker_tokens(self.multi_input.toPlainText())
            if not tickers:
                raise ValueError("Provide tickers for multiple stock search.")
            self._append_chat(f"Search mode: multiple stocks | count={len(tickers)}")
            self._append_chat(f"Tickers: {', '.join(tickers)}")
            cmd += ["--multi"] + tickers

        else:
            path = self.file_input.text().strip()
            if not path:
                raise ValueError("Choose a ticker file.")
            self._append_chat(f"Search mode: file | {path}")
            cmd += ["--ticker-file", path]

        return cmd

    def _run_analysis(self) -> None:
        if self.proc and self.proc.poll() is None:
            QMessageBox.information(self, "Analyzer", "Analysis is already running.")
            return

        try:
            cmd = self._build_cmd()
        except ValueError as e:
            QMessageBox.critical(self, "Input Error", str(e))
            return

        self.status_label.setText("Running...")
        self.run_btn.setEnabled(False)
        self.summary_btn.setEnabled(False)
        self.last_report_path = None
        self.run_started_at = datetime.now()
        self._append_chat(f"[{datetime.now().strftime('%H:%M:%S')}] Starting analysis...")

        def worker() -> None:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            report_path = None
            try:
                self.proc = subprocess.Popen(
                    cmd,
                    cwd=str(PROJECT_ROOT),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env=env,
                )
                assert self.proc.stdout is not None
                for line in self.proc.stdout:
                    if "Report saved:" in line:
                        maybe = line.split("Report saved:", 1)[1].strip().strip('"').strip("'")
                        if maybe:
                            report_path = maybe
                    self.event_queue.put(("log", line.rstrip("\n")))
                rc = self.proc.wait()
                self.event_queue.put(("done", (rc, report_path)))
            except Exception as ex:
                self.event_queue.put(("log", f"[GUI] Failed to run analyzer: {ex}"))
                self.event_queue.put(("done", (1, report_path)))

        threading.Thread(target=worker, daemon=True).start()

    def _drain_events(self) -> None:
        while True:
            try:
                kind, payload = self.event_queue.get_nowait()
            except queue.Empty:
                break

            if kind == "log":
                self._append_chat(str(payload))
                continue

            if kind == "done":
                if isinstance(payload, tuple):
                    code = int(payload[0])
                    maybe_path = payload[1]
                else:
                    code = int(payload)
                    maybe_path = None
                self.run_btn.setEnabled(True)
                self.status_label.setText("Idle")
                if code == 0:
                    self._append_chat("[GUI] Analysis completed successfully.")
                    resolved: Optional[Path] = None
                    if maybe_path:
                        p = Path(str(maybe_path))
                        if p.exists():
                            resolved = p
                    if resolved is None:
                        resolved = self._find_latest_report(self.run_started_at)
                    if resolved and resolved.exists():
                        self.last_report_path = resolved
                        self.summary_btn.setEnabled(True)
                        self._append_chat(f"[GUI] Summary ready. Click 'View Analysis Summary'.")
                    else:
                        self._append_chat("[GUI] Analysis finished, but report file could not be located.")
                else:
                    self._append_chat(f"[GUI] Analysis failed with exit code {code}.")
                continue

            if kind == "profile":
                job_id, profile = payload
                if job_id != self.profile_job_id:
                    continue
                self._apply_profile(profile)
                continue

            if kind == "candles":
                job_id, symbol, rng, interval, candles = payload
                if job_id != self.chart_job_id:
                    continue
                self._render_candles(symbol, rng, interval, candles)

    def _apply_profile(self, profile: dict) -> None:
        symbol = profile.get("symbol") or "-"
        company = profile.get("company") or "N/A"
        sector = profile.get("sector") or "N/A"
        price = profile.get("price")
        mcap = profile.get("market_cap")
        pe = profile.get("pe")
        website = profile.get("website")
        history = profile.get("history") or "No company history available."
        qoq = profile.get("qoq_growth")
        yoy = profile.get("yoy_growth")
        earnings_date = profile.get("nearest_earnings_date")
        chart_points = profile.get("chart_points") or []
        chart_dates = profile.get("chart_dates") or []
        quarter_labels = profile.get("quarter_labels") or []
        quarterly_results = profile.get("quarterly_results") or {}
        pix = profile.get("logo_pixmap")

        self.profile_symbol.setText(f"Symbol: {symbol}")
        self.profile_company.setText(f"Company: {company}")
        self.profile_sector.setText(f"Sector: {sector}")
        if website:
            self.profile_website.setText(f'Website: <a href="{website}">{website}</a>')
        else:
            self.profile_website.setText("Website: N/A")
        self.profile_price.setText(f"Current Price: {compact_currency(price) if price is not None else 'N/A'}")
        self.profile_mcap.setText(f"Market Cap: {compact_currency(mcap) if mcap is not None else 'N/A'}")
        self.profile_pe.setText(f"P/E: {pe:.2f}" if isinstance(pe, (int, float)) else "P/E: N/A")
        self.profile_earnings_date.setText(
            f"Nearest Earnings Date: {earnings_date if earnings_date else 'N/A'}"
        )
        self.profile_qoq.setText(
            f"QoQ Earnings Growth: {qoq*100:.2f}%" if isinstance(qoq, (int, float)) else "QoQ Earnings Growth: N/A"
        )
        self.profile_yoy.setText(
            f"YoY Earnings Growth: {yoy*100:.2f}%" if isinstance(yoy, (int, float)) else "YoY Earnings Growth: N/A"
        )
        self.profile_history.setPlainText(history)
        self._populate_quarterly_table(quarter_labels, quarterly_results)

        if isinstance(pix, QPixmap):
            self.logo_label.setPixmap(pix)
        else:
            self.logo_label.setPixmap(self._build_placeholder_logo(symbol))

        self.profile_status.setText("Status: Profile ready")
        self._append_price_chart_to_chat(symbol, chart_dates, chart_points)

    def _append_price_chart_to_chat(self, symbol: str, dates: list[str], closes: list[float]) -> None:
        if not closes:
            self._append_chat(f"[Chart] {symbol} 1M price movement: N/A")
            return
        start = closes[0]
        end = closes[-1]
        high = max(closes)
        low = min(closes)
        change = ((end / start) - 1) * 100 if start else 0
        chart = sparkline(closes)
        start_date = dates[0] if dates else "-"
        end_date = dates[-1] if dates else "-"
        self._append_chat(f"[Chart] {symbol} 1M ({start_date} to {end_date})")
        self._append_chat(f"[Chart] {chart}")
        self._append_chat(
            f"[Chart] Start {start:.2f} | End {end:.2f} | High {high:.2f} | Low {low:.2f} | Change {change:+.2f}%"
        )

    def _populate_quarterly_table(self, quarter_labels: list[str], quarterly_results: dict[str, dict[str, float]]) -> None:
        if not quarter_labels or not quarterly_results:
            self.quarterly_table.setRowCount(1)
            self.quarterly_table.setColumnCount(1)
            self.quarterly_table.setHorizontalHeaderLabels(["Quarterly Results"])
            self.quarterly_table.setItem(0, 0, QTableWidgetItem("No quarterly data available for the selected stock."))
            return

        metrics = sorted(quarterly_results.keys())
        self.quarterly_table.setRowCount(len(metrics))
        self.quarterly_table.setColumnCount(1 + len(quarter_labels))
        headers = ["Metric"] + quarter_labels
        self.quarterly_table.setHorizontalHeaderLabels(headers)

        for r, metric in enumerate(metrics):
            self.quarterly_table.setItem(r, 0, QTableWidgetItem(metric))
            vals = quarterly_results.get(metric, {})
            for c, q in enumerate(quarter_labels, start=1):
                raw = vals.get(q)
                txt = fmt_num(raw) if isinstance(raw, (int, float)) else "N/A"
                self.quarterly_table.setItem(r, c, QTableWidgetItem(txt))

        self.quarterly_table.horizontalHeader().setStretchLastSection(True)

    def _show_summary_dialog(self) -> None:
        if not self.last_report_path or not self.last_report_path.exists():
            QMessageBox.information(self, "Summary", "No completed analysis summary is available yet.")
            return
        try:
            summary = extract_summary_from_workbook(self.last_report_path)
            dlg = SummaryDialog(summary, self.last_report_path, self)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, "Summary Error", f"Could not parse report summary:\n{e}")
            return

    def _append_chat(self, text: str) -> None:
        self.chat_box.appendPlainText(text)


def main() -> None:
    app = QApplication(sys.argv)
    window = AnalyzerWindow()
    window.showFullScreen()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
