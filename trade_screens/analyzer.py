#!/usr/bin/env python3
"""
==========================================================================
NSE STOCK INVESTMENT ANALYZER
==========================================================================
Accepts any NSE-listed stock ticker(s), fetches real market data,
runs 7 quantitative models, and generates a professional Excel workbook.

Usage:
    python nse_stock_analyzer.py RELIANCE
    python nse_stock_analyzer.py HAL BEL TATAPOWER DIXON
    python nse_stock_analyzer.py --interactive
    python nse_stock_analyzer.py --ticker-file data/nse_equity_symbols_eq.txt
    python nse_stock_analyzer.py --single RELIANCE
    python nse_stock_analyzer.py --multi HAL BEL RELIANCE
    python nse_stock_analyzer.py --screener --ticker-file data/nse_equity_symbols_eq.txt --top-n 10

Models included:
    1. DCF Valuation with Monte Carlo
    2. Fama-French 5-Factor Model
    3. Black-Litterman Portfolio Optimization
    4. GARCH(1,1) Volatility Forecasting
    5. Kelly Criterion & VaR Position Sizing
    6. ESG Composite Scoring
    7. Technical Momentum & Mean Reversion Signals

Dependencies:
    pip install yfinance pandas numpy scipy openpyxl
==========================================================================
"""

import sys
import argparse
import json
import math
import os
import warnings
from datetime import datetime, timedelta
from pathlib import Path
import requests

import numpy as np
import pandas as pd
import yfinance as yf
from scipy.optimize import minimize
from scipy.stats import norm

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

PROJECT_ROOT = Path(__file__).resolve().parent.parent
REPORTS_DIR = PROJECT_ROOT / "reports"

# ============================================================
# STYLING CONSTANTS
# ============================================================
BLUE_FONT = Font(name='Arial', color='0000FF', size=10)
BLACK_FONT = Font(name='Arial', color='000000', size=10)
BLACK_BOLD = Font(name='Arial', color='000000', size=10, bold=True)
GREEN_FONT = Font(name='Arial', color='008000', size=10)
HEADER_FONT = Font(name='Arial', color='FFFFFF', size=11, bold=True)
TITLE_FONT = Font(name='Arial', color='000000', size=14, bold=True)
SECTION_FONT = Font(name='Arial', color='1F4E79', size=11, bold=True)
SUBTITLE_FONT = Font(name='Arial', size=10, italic=True, color='555555')
SMALL_GRAY = Font(name='Arial', size=9, color='888888', italic=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')
RESULT_FILL = PatternFill('solid', fgColor='D6EAF8')
GREEN_FILL = PatternFill('solid', fgColor='D5F5E3')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
WARNING_FILL = PatternFill('solid', fgColor='FFFF00')
RED_FILL = PatternFill('solid', fgColor='FADBD8')

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrapText=True)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

PCT_FMT = '0.0%'
CUR_FMT = 'â‚¹#,##0'
CUR_DEC = 'â‚¹#,##0.00'
NUM_FMT = '#,##0'
DEC_FMT = '#,##0.00'
MULT_FMT = '0.0"x"'


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_cell(ws, row, col, font=None, fill=None, fmt=None, align=None):
    cell = ws.cell(row=row, column=col)
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    cell.border = THIN_BORDER
    return cell

def auto_width(ws, min_w=14, max_w=30):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min_w


# ============================================================
# DATA FETCHER
# ============================================================
class NSEDataFetcher:
    """Fetches and processes data for NSE-listed stocks."""

    def __init__(self, tickers: list, benchmark='^NSEI'):
        self.raw_tickers = tickers
        self.tickers = [t.upper().replace('.NS', '') + '.NS' for t in tickers]
        self.benchmark = benchmark  # ^NSEI = NIFTY 50
        self.data = {}
        self.prices = None
        self.returns = None
        self.benchmark_data = None
        self.http = requests.Session()
        self.http.trust_env = False
        self.http.headers.update({'User-Agent': 'Mozilla/5.0'})
    
    def _clear_broken_proxy_env(self):
        # Some local setups set a dead loopback proxy (127.0.0.1:9),
        # which makes Yahoo requests always fail.
        proxy_keys = [
            'HTTP_PROXY', 'HTTPS_PROXY', 'ALL_PROXY',
            'http_proxy', 'https_proxy', 'all_proxy'
        ]
        bad_tokens = ('127.0.0.1:9', 'localhost:9')
        cleared = []
        for key in proxy_keys:
            val = os.environ.get(key, '')
            if any(tok in val for tok in bad_tokens):
                os.environ.pop(key, None)
                cleared.append(key)
        if cleared:
            print(f"  â„¹ Cleared invalid proxy settings: {', '.join(cleared)}")

    def _fetch_chart(self, symbol: str, rng: str = '3y', interval: str = '1d'):
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
        try:
            resp = self.http.get(url, params={'range': rng, 'interval': interval}, timeout=20)
            if resp.status_code != 200:
                return pd.Series(dtype=float), {}
            payload = resp.json()
            result = payload.get('chart', {}).get('result')
            if not result:
                return pd.Series(dtype=float), {}
            node = result[0]
            meta = node.get('meta', {}) or {}
            ts = node.get('timestamp') or []
            closes = (((node.get('indicators') or {}).get('quote') or [{}])[0].get('close') or [])
            if not ts or not closes:
                return pd.Series(dtype=float), meta
            idx = pd.to_datetime(ts, unit='s')
            series = pd.Series(closes, index=idx).dropna().astype(float)
            return series, meta
        except Exception:
            return pd.Series(dtype=float), {}

    def _search_symbol_profile(self, symbol: str):
        try:
            resp = self.http.get(
                "https://query2.finance.yahoo.com/v1/finance/search",
                params={'q': symbol, 'quotesCount': 1, 'newsCount': 0},
                timeout=20
            )
            if resp.status_code != 200:
                return {}
            quotes = resp.json().get('quotes') or []
            if not quotes:
                return {}
            q = quotes[0]
            return {
                'name': q.get('longname') or q.get('shortname') or symbol.replace('.NS', ''),
                'sector': q.get('sectorDisp') or q.get('sector') or 'N/A',
                'industry': q.get('industryDisp') or q.get('industry') or 'N/A',
            }
        except Exception:
            return {}

    def _extract_latest_ts_value(self, entry: dict, key: str):
        items = entry.get(key) or []
        if not items:
            return 0
        values = [x.get('reportedValue', {}).get('raw') for x in items if isinstance(x, dict)]
        values = [v for v in values if isinstance(v, (int, float))]
        return values[-1] if values else 0

    def _extract_ttm_and_yoy(self, entry: dict, key: str):
        items = entry.get(key) or []
        vals = [x.get('reportedValue', {}).get('raw') for x in items if isinstance(x, dict)]
        vals = [v for v in vals if isinstance(v, (int, float))]
        if not vals:
            return 0, 0
        ttm = sum(vals[-4:]) if len(vals) >= 4 else vals[-1]
        yoy = 0
        if len(vals) >= 5 and vals[-5] != 0:
            yoy = vals[-1] / vals[-5] - 1
        return ttm, yoy

    def _fetch_timeseries_fundamentals(self, symbol: str):
        types = [
            'quarterlyTotalRevenue', 'quarterlyEBITDA', 'quarterlyNetIncome',
            'quarterlyBasicAverageShares', 'quarterlyFreeCashFlow',
            'quarterlyTotalDebt', 'quarterlyCashAndCashEquivalents',
            'quarterlyStockholdersEquity',
        ]
        try:
            resp = self.http.get(
                f"https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/{symbol}",
                params={
                    'type': ','.join(types),
                    'merge': 'false',
                    'period1': '1609459200',   # 2021-01-01
                    'period2': str(int(datetime.now().timestamp())),
                },
                timeout=20
            )
            if resp.status_code != 200:
                return {}
            res = resp.json().get('timeseries', {}).get('result') or []
            entry = {}
            for item in res:
                typ = ((item.get('meta') or {}).get('type') or [None])[0]
                if typ:
                    entry[typ] = item.get(typ, [])

            revenue_ttm, rev_yoy = self._extract_ttm_and_yoy(entry, 'quarterlyTotalRevenue')
            ebitda_ttm, _ = self._extract_ttm_and_yoy(entry, 'quarterlyEBITDA')
            net_income_ttm, ni_yoy = self._extract_ttm_and_yoy(entry, 'quarterlyNetIncome')
            fcf_ttm, _ = self._extract_ttm_and_yoy(entry, 'quarterlyFreeCashFlow')

            return {
                'revenue': revenue_ttm,
                'ebitda': ebitda_ttm,
                'net_income': net_income_ttm,
                'free_cashflow': fcf_ttm,
                'shares_outstanding': self._extract_latest_ts_value(entry, 'quarterlyBasicAverageShares'),
                'total_debt': self._extract_latest_ts_value(entry, 'quarterlyTotalDebt'),
                'total_cash': self._extract_latest_ts_value(entry, 'quarterlyCashAndCashEquivalents'),
                'stockholders_equity': self._extract_latest_ts_value(entry, 'quarterlyStockholdersEquity'),
                'revenue_growth': rev_yoy,
                'earnings_growth': ni_yoy,
            }
        except Exception:
            return {}

    def _build_fallback_record(self, ticker: str):
        profile = self._search_symbol_profile(ticker)
        ts_fb = self._fetch_timeseries_fundamentals(ticker)
        series, meta = self._fetch_chart(ticker, rng='3y', interval='1d')
        if series.empty:
            return self._empty_record(ticker)

        current_price = float(meta.get('regularMarketPrice') or series.iloc[-1] or 0)
        high_52 = float(meta.get('fiftyTwoWeekHigh') or (series.tail(252).max() if len(series) >= 252 else series.max()))
        low_52 = float(meta.get('fiftyTwoWeekLow') or (series.tail(252).min() if len(series) >= 252 else series.min()))
        sma50 = float(series.tail(50).mean()) if len(series) >= 50 else 0
        sma200 = float(series.tail(200).mean()) if len(series) >= 200 else 0

        shares = ts_fb.get('shares_outstanding') or 0
        revenue = ts_fb.get('revenue') or 0
        ebitda = ts_fb.get('ebitda') or 0
        net_income = ts_fb.get('net_income') or 0
        total_debt = ts_fb.get('total_debt') or 0
        total_cash = ts_fb.get('total_cash') or 0
        equity = ts_fb.get('stockholders_equity') or 0
        free_cashflow = ts_fb.get('free_cashflow') or 0

        market_cap = current_price * shares if current_price and shares else 0
        book_value = (equity / shares) if equity and shares else 0
        pb_ratio = (current_price / book_value) if current_price and book_value else 0
        enterprise_value = market_cap + total_debt - total_cash if market_cap else 0
        pe_ratio = (market_cap / net_income) if market_cap and net_income > 0 else 0
        ev_ebitda = (enterprise_value / ebitda) if enterprise_value and ebitda > 0 else 0
        roe = (net_income / equity) if net_income and equity else 0
        profit_margin = (net_income / revenue) if net_income and revenue else 0
        operating_margin = (ebitda / revenue) if ebitda and revenue else 0
        debt_to_equity = (total_debt / equity) * 100 if total_debt and equity else 0

        record = self._empty_record(ticker)
        record.update({
            'name': profile.get('name') or meta.get('longName') or meta.get('shortName') or ticker.replace('.NS', ''),
            'sector': profile.get('sector') or 'N/A',
            'industry': profile.get('industry') or 'N/A',
            'market_cap': market_cap,
            'current_price': current_price,
            'pe_ratio': pe_ratio,
            'pb_ratio': pb_ratio,
            'ev_ebitda': ev_ebitda,
            'revenue': revenue,
            'ebitda': ebitda,
            'net_income': net_income,
            'total_debt': total_debt,
            'total_cash': total_cash,
            'shares_outstanding': shares,
            'roe': roe,
            'profit_margin': profit_margin,
            'operating_margin': operating_margin,
            'gross_margin': operating_margin,
            'revenue_growth': ts_fb.get('revenue_growth') or 0,
            'earnings_growth': ts_fb.get('earnings_growth') or 0,
            'fifty_two_week_high': high_52,
            'fifty_two_week_low': low_52,
            'two_hundred_day_avg': sma200,
            'fifty_day_avg': sma50,
            'book_value': book_value,
            'debt_to_equity': debt_to_equity,
            'free_cashflow': free_cashflow,
            'enterprise_value': enterprise_value,
        })
        return record

    def fetch_all(self):
        print(f"\n{'='*60}")
        print(f"  FETCHING DATA FOR {len(self.tickers)} STOCKS")
        print(f"{'='*60}")

        self._clear_broken_proxy_env()

        # Keep yfinance timezone cache in a writable local folder.
        try:
            cache_dir = PROJECT_ROOT / ".yf_cache"
            cache_dir.mkdir(parents=True, exist_ok=True)
            yf.set_tz_cache_location(str(cache_dir))
        except Exception:
            pass

        all_tickers = self.tickers + [self.benchmark]
        end = datetime.now()
        start = end - timedelta(days=756)  # ~3 years

        print(f"  Downloading price history (3 years)...")
        prices = yf.download(
            all_tickers,
            start=start,
            end=end,
            progress=False,
            auto_adjust=False,
            group_by='column',
            threads=False,
            session=self.http,
        )

        # Normalize close prices to a DataFrame keyed by ticker
        close = pd.DataFrame()
        if isinstance(prices.columns, pd.MultiIndex):
            if 'Close' in prices.columns.get_level_values(0):
                close = prices['Close'].copy()
            elif 'Adj Close' in prices.columns.get_level_values(0):
                close = prices['Adj Close'].copy()
        else:
            if 'Close' in prices.columns:
                close = prices[['Close']].copy()
                close.columns = [self.tickers[0]]
            elif 'Adj Close' in prices.columns:
                close = prices[['Adj Close']].copy()
                close.columns = [self.tickers[0]]

        if close.empty:
            print("  âš  Warning: Could not download bulk close-price data. Retrying per ticker...")
            close = pd.DataFrame()

        # Retry missing symbols one-by-one (helps when bulk download partially fails)
        missing = [t for t in all_tickers if t not in close.columns]
        for t in missing:
            try:
                single = yf.download(
                    t,
                    start=start,
                    end=end,
                    progress=False,
                    auto_adjust=False,
                    group_by='column',
                    threads=False,
                    session=self.http,
                )
                if isinstance(single.columns, pd.MultiIndex):
                    if 'Close' in single.columns.get_level_values(0):
                        s = single['Close'].squeeze()
                    elif 'Adj Close' in single.columns.get_level_values(0):
                        s = single['Adj Close'].squeeze()
                    else:
                        s = pd.Series(dtype=float)
                else:
                    if 'Close' in single.columns:
                        s = single['Close']
                    elif 'Adj Close' in single.columns:
                        s = single['Adj Close']
                    else:
                        s = pd.Series(dtype=float)

                if not s.empty:
                    close[t] = s
                else:
                    # Fallback via Yahoo chart endpoint (no crumb required).
                    s2, _ = self._fetch_chart(t, rng='3y', interval='1d')
                    if not s2.empty:
                        close[t] = s2
                    else:
                        print(f"  âš  Warning: No price history available for {t}")
            except Exception as e:
                print(f"  âš  Warning: Per-ticker price retry failed for {t}: {e}")

        self.prices = close.sort_index()
        self.returns = self.prices.pct_change().dropna() if not self.prices.empty else pd.DataFrame()

        # Fetch individual stock info
        for ticker in self.tickers:
            print(f"  Fetching fundamentals: {ticker}...")
            try:
                stock = yf.Ticker(ticker, session=self.http)
                info = stock.info if isinstance(stock.info, dict) else {}
                fast_info = {}
                try:
                    fast_info = dict(stock.fast_info)
                except Exception:
                    fast_info = {}

                profile_fb = self._search_symbol_profile(ticker)
                ts_fb = self._fetch_timeseries_fundamentals(ticker)

                px = self.prices[ticker].dropna() if ticker in self.prices.columns else pd.Series(dtype=float)
                hist_last = float(px.iloc[-1]) if not px.empty else 0
                hist_high = float(px.max()) if not px.empty else 0
                hist_low = float(px.min()) if not px.empty else 0
                hist_sma50 = float(px.tail(50).mean()) if len(px) >= 50 else 0
                hist_sma200 = float(px.tail(200).mean()) if len(px) >= 200 else 0

                chart_series, chart_meta = self._fetch_chart(ticker, rng='3y', interval='1d')
                if not chart_series.empty:
                    if px.empty:
                        px = chart_series
                    hist_last = float(px.iloc[-1])
                    hist_high = float(px.tail(252).max()) if len(px) >= 252 else float(px.max())
                    hist_low = float(px.tail(252).min()) if len(px) >= 252 else float(px.min())
                    hist_sma50 = float(px.tail(50).mean()) if len(px) >= 50 else hist_sma50
                    hist_sma200 = float(px.tail(200).mean()) if len(px) >= 200 else hist_sma200

                current_price = (
                    info.get('currentPrice')
                    or info.get('regularMarketPrice')
                    or fast_info.get('lastPrice')
                    or chart_meta.get('regularMarketPrice')
                    or hist_last
                )
                shares_outstanding = info.get('sharesOutstanding') or ts_fb.get('shares_outstanding') or 0
                market_cap = info.get('marketCap') or fast_info.get('marketCap') or (current_price * shares_outstanding if shares_outstanding and current_price else 0)
                fifty_two_week_high = info.get('fiftyTwoWeekHigh') or fast_info.get('yearHigh') or chart_meta.get('fiftyTwoWeekHigh') or hist_high
                fifty_two_week_low = info.get('fiftyTwoWeekLow') or fast_info.get('yearLow') or chart_meta.get('fiftyTwoWeekLow') or hist_low
                fifty_day_avg = info.get('fiftyDayAverage') or hist_sma50
                two_hundred_day_avg = info.get('twoHundredDayAverage') or hist_sma200

                revenue = info.get('totalRevenue') or ts_fb.get('revenue') or 0
                ebitda = info.get('ebitda') or ts_fb.get('ebitda') or 0
                net_income = info.get('netIncomeToCommon') or ts_fb.get('net_income') or 0
                total_debt = info.get('totalDebt') or ts_fb.get('total_debt') or 0
                total_cash = info.get('totalCash') or ts_fb.get('total_cash') or 0
                free_cashflow = info.get('freeCashflow') or ts_fb.get('free_cashflow') or 0
                equity = ts_fb.get('stockholders_equity') or 0
                book_value = info.get('bookValue') or ((equity / shares_outstanding) if equity and shares_outstanding else 0)
                pb_ratio = info.get('priceToBook') or ((current_price / book_value) if book_value else 0)
                enterprise_value = info.get('enterpriseValue') or (market_cap + total_debt - total_cash if market_cap else 0)
                pe_ratio = info.get('trailingPE') or ((market_cap / net_income) if market_cap and net_income > 0 else 0)
                ev_ebitda = info.get('enterpriseToEbitda') or ((enterprise_value / ebitda) if enterprise_value and ebitda > 0 else 0)
                roe = info.get('returnOnEquity') or ((net_income / equity) if net_income and equity else 0)
                profit_margin = info.get('profitMargins') or ((net_income / revenue) if net_income and revenue else 0)
                operating_margin = info.get('operatingMargins') or ((ebitda / revenue) if ebitda and revenue else 0)
                gross_margin = info.get('grossMargins') or operating_margin
                debt_to_equity = info.get('debtToEquity') or ((total_debt / equity) * 100 if total_debt and equity else 0)
                revenue_growth = info.get('revenueGrowth') or ts_fb.get('revenue_growth') or 0
                earnings_growth = info.get('earningsGrowth') or ts_fb.get('earnings_growth') or 0

                self.data[ticker] = {
                    'info': info,
                    'name': info.get('longName') or info.get('shortName') or profile_fb.get('name') or chart_meta.get('longName') or chart_meta.get('shortName') or ticker.replace('.NS', ''),
                    'sector': info.get('sector') or profile_fb.get('sector') or 'N/A',
                    'industry': info.get('industry') or profile_fb.get('industry') or 'N/A',
                    'market_cap': market_cap,
                    'current_price': current_price,
                    'pe_ratio': pe_ratio,
                    'forward_pe': info.get('forwardPE', 0),
                    'pb_ratio': pb_ratio,
                    'ev_ebitda': ev_ebitda,
                    'revenue': revenue,
                    'ebitda': ebitda,
                    'net_income': net_income,
                    'total_debt': total_debt,
                    'total_cash': total_cash,
                    'shares_outstanding': shares_outstanding,
                    'dividend_yield': info.get('dividendYield', 0),
                    'roe': roe,
                    'profit_margin': profit_margin,
                    'operating_margin': operating_margin,
                    'gross_margin': gross_margin,
                    'revenue_growth': revenue_growth,
                    'earnings_growth': earnings_growth,
                    'beta': info.get('beta', 1.0),
                    'fifty_two_week_high': fifty_two_week_high,
                    'fifty_two_week_low': fifty_two_week_low,
                    'two_hundred_day_avg': two_hundred_day_avg,
                    'fifty_day_avg': fifty_day_avg,
                    'book_value': book_value,
                    'debt_to_equity': debt_to_equity,
                    'free_cashflow': free_cashflow,
                    'enterprise_value': enterprise_value,
                }
            except Exception as e:
                print(f"  âš  Warning: Could not fetch full data for {ticker}: {e}")
                self.data[ticker] = self._build_fallback_record(ticker)

        # Fetch benchmark info
        print(f"  Fetching benchmark: NIFTY 50...")
        try:
            bench = yf.Ticker(self.benchmark, session=self.http)
            self.benchmark_data = bench.info
        except:
            self.benchmark_data = {}

        print(f"\n  âœ… Data fetch complete for {len(self.data)} stocks")
        return self

    def _empty_record(self, ticker):
        record = {k: 0 for k in ['market_cap', 'current_price', 'pe_ratio', 'forward_pe',
                                'pb_ratio', 'ev_ebitda', 'revenue', 'ebitda', 'net_income',
                                'total_debt', 'total_cash', 'shares_outstanding', 'dividend_yield',
                                'roe', 'profit_margin', 'operating_margin', 'gross_margin',
                                'revenue_growth', 'earnings_growth', 'beta',
                                'fifty_two_week_high', 'fifty_two_week_low',
                                'two_hundred_day_avg', 'fifty_day_avg', 'book_value',
                                'debt_to_equity', 'free_cashflow', 'enterprise_value']}
        record.update({
            'name': ticker.replace('.NS', ''),
            'sector': 'N/A',
            'industry': 'N/A',
            'beta': 1.0,
            'info': {},
        })
        return record

    def get_returns_matrix(self):
        if self.returns is None or self.returns.empty:
            return pd.DataFrame()
        cols = [t for t in self.tickers if t in self.returns.columns]
        return self.returns[cols].dropna()

    def get_covariance(self):
        ret = self.get_returns_matrix()
        if ret.empty:
            return pd.DataFrame()
        return ret.cov() * 252  # annualized

    def get_correlation(self):
        ret = self.get_returns_matrix()
        if ret.empty:
            return pd.DataFrame()
        return ret.corr()

    def get_volatilities(self):
        ret = self.get_returns_matrix()
        if ret.empty:
            return {}
        return (ret.std() * np.sqrt(252)).to_dict()

    def get_annualized_returns(self):
        ret = self.get_returns_matrix()
        if ret.empty:
            return {}
        total_days = len(ret)
        total_return = (1 + ret).prod()
        ann = total_return ** (252 / total_days) - 1
        return ann.to_dict()


# ============================================================
# QUANTITATIVE MODELS
# ============================================================
class QuantModels:
    """All quantitative models for investment analysis."""

    def __init__(self, fetcher: NSEDataFetcher):
        self.f = fetcher
        self.rf = 0.071  # India 10Y

    # --- GARCH(1,1) ---
    def fit_garch(self, ticker):
        if ticker not in self.f.returns.columns:
            return {'omega': 1e-5, 'alpha': 0.08, 'beta': 0.90}
        ret = self.f.returns[ticker].dropna().values
        # Simple method-of-moments GARCH estimation
        var_series = ret ** 2
        mean_var = np.mean(var_series)
        # Estimate via autocorrelation of squared returns
        if len(var_series) > 10:
            autocorr1 = np.corrcoef(var_series[:-1], var_series[1:])[0, 1]
            autocorr1 = max(0.01, min(autocorr1, 0.99))
        else:
            autocorr1 = 0.9
        beta = max(0.5, min(autocorr1, 0.95))
        alpha = max(0.01, min(0.15, 1 - beta - 0.02))
        omega = mean_var * (1 - alpha - beta)
        omega = max(1e-7, omega)
        return {'omega': omega, 'alpha': alpha, 'beta': beta}

    def garch_forecast(self, params, current_var, n_days=252):
        omega, alpha, beta = params['omega'], params['alpha'], params['beta']
        lr_var = omega / max(1 - alpha - beta, 0.001)
        forecasts = []
        for h in range(1, n_days + 1):
            fv = lr_var + (alpha + beta)**h * (current_var - lr_var)
            forecasts.append(fv)
        return forecasts

    # --- FACTOR MODEL ---
    def estimate_factor_betas(self, ticker):
        if ticker not in self.f.returns.columns or self.f.benchmark not in self.f.returns.columns:
            return {'market': 1.0, 'smb': 0.0, 'hml': 0.0, 'rmw': 0.0, 'cma': 0.0}
        stock_ret = self.f.returns[ticker].dropna()
        bench_ret = self.f.returns[self.f.benchmark].reindex(stock_ret.index).dropna()
        common = stock_ret.index.intersection(bench_ret.index)
        if len(common) < 30:
            return {'market': 1.0, 'smb': 0.0, 'hml': 0.0, 'rmw': 0.0, 'cma': 0.0}
        s, b = stock_ret.loc[common].values, bench_ret.loc[common].values
        cov_sb = np.cov(s, b)[0, 1]
        var_b = np.var(b)
        market_beta = cov_sb / var_b if var_b > 0 else 1.0
        # Heuristic factor betas based on stock characteristics
        info = self.f.data.get(ticker, {})
        mcap = info.get('market_cap', 0)
        pb = info.get('pb_ratio', 0)
        roe_val = info.get('roe', 0)
        # Size: small caps get positive SMB beta
        smb = max(-0.5, min(0.8, 0.5 - (mcap / 1e12) * 0.3)) if mcap > 0 else 0.2
        # Value: high book-to-market gets positive HML beta
        hml = max(-0.5, min(0.8, (1 / max(pb, 0.5) - 0.5) * 0.5)) if pb > 0 else 0.2
        # Profitability: high ROE gets positive RMW beta
        rmw = max(-0.3, min(0.8, (roe_val if roe_val else 0.1) * 2)) if roe_val else 0.2
        # Investment: conservative (low capex/assets) gets positive CMA
        cma = 0.1  # Default neutral
        return {'market': round(market_beta, 3), 'smb': round(smb, 2),
                'hml': round(hml, 2), 'rmw': round(rmw, 2), 'cma': round(cma, 2)}

    def factor_expected_return(self, betas):
        premiums = {'market': 0.059, 'smb': 0.025, 'hml': 0.035, 'rmw': 0.030, 'cma': 0.020}
        er = self.rf
        for f, b in betas.items():
            er += b * premiums.get(f, 0)
        return er

    # --- BLACK-LITTERMAN ---
    def black_litterman(self, views=None, confidence=None):
        tickers = [t for t in self.f.tickers if t in self.f.returns.columns]
        n = len(tickers)
        if n == 0:
            return {}

        cov = self.f.get_covariance()
        if cov.empty:
            return {}
        if any(t not in cov.index for t in tickers):
            return {}
        cov_mat = cov.loc[tickers, tickers].values
        # Ensure positive definite
        eigvals = np.linalg.eigvalsh(cov_mat)
        if np.any(eigvals <= 0):
            cov_mat += np.eye(n) * 1e-6

        mcaps = np.array([self.f.data.get(t, {}).get('market_cap', 1e9) for t in tickers], dtype=float)
        mcap_sum = mcaps.sum()
        if mcap_sum <= 0:
            mcap_w = np.ones(n) / n
        else:
            mcap_w = mcaps / mcap_sum

        delta = 2.5
        tau = 0.05
        pi = delta * cov_mat @ mcap_w  # Equilibrium returns

        if views and len(views) > 0:
            k = len(views)
            P = np.zeros((k, n))
            Q = np.zeros(k)
            omega_diag = np.zeros(k)
            for i, v in enumerate(views):
                idx = tickers.index(v['ticker']) if v['ticker'] in tickers else -1
                if idx >= 0:
                    P[i, idx] = 1.0
                    Q[i] = v['return']
                    omega_diag[i] = 1 / max(v.get('confidence', 0.5), 0.1) * 0.01
            omega = np.diag(omega_diag)
            try:
                tau_cov_inv = np.linalg.inv(tau * cov_mat)
                omega_inv = np.linalg.inv(omega)
                bl_precision = tau_cov_inv + P.T @ omega_inv @ P
                bl_returns = np.linalg.inv(bl_precision) @ (tau_cov_inv @ pi + P.T @ omega_inv @ Q)
            except:
                bl_returns = pi
        else:
            bl_returns = pi

        # Optimize for max Sharpe
        def neg_sharpe(w):
            pr = w @ bl_returns
            pv = np.sqrt(w @ cov_mat @ w)
            return -(pr - self.rf) / max(pv, 1e-8)

        constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]
        bounds = [(0.02, 0.40)] * n
        x0 = mcap_w.copy()
        result = minimize(neg_sharpe, x0, method='SLSQP', bounds=bounds, constraints=constraints)
        opt_w = result.x if result.success else mcap_w

        port_ret = opt_w @ bl_returns
        port_vol = np.sqrt(opt_w @ cov_mat @ opt_w)
        sharpe = (port_ret - self.rf) / max(port_vol, 1e-8)

        return {
            'tickers': tickers,
            'equilibrium': dict(zip(tickers, pi)),
            'bl_returns': dict(zip(tickers, bl_returns)),
            'mcap_weights': dict(zip(tickers, mcap_w)),
            'optimal_weights': dict(zip(tickers, opt_w)),
            'port_return': port_ret,
            'port_vol': port_vol,
            'sharpe': sharpe,
        }

    # --- DCF ---
    def dcf_valuation(self, ticker, scenarios=None):
        info = self.f.data.get(ticker, {})
        revenue = info.get('revenue', 0) or 0
        ebitda = info.get('ebitda', 0) or 0
        fcf = info.get('free_cashflow', 0) or 0
        debt = info.get('total_debt', 0) or 0
        cash = info.get('total_cash', 0) or 0
        shares = info.get('shares_outstanding', 0) or 1
        price = info.get('current_price', 0) or 0
        rev_growth = info.get('revenue_growth', 0) or 0.10
        ebitda_margin = (ebitda / revenue) if revenue > 0 else 0.15

        # Convert to Crores
        rev_cr = revenue / 1e7
        net_debt_cr = (debt - cash) / 1e7
        shares_cr = shares / 1e7

        if not scenarios:
            scenarios = {
                'Base': {'g1': rev_growth, 'g2': rev_growth * 0.8, 'margin': ebitda_margin,
                         'wacc': 0.11, 'tg': 0.05},
                'Bull': {'g1': rev_growth * 1.3, 'g2': rev_growth * 1.1, 'margin': ebitda_margin * 1.1,
                         'wacc': 0.10, 'tg': 0.06},
                'Bear': {'g1': rev_growth * 0.6, 'g2': rev_growth * 0.5, 'margin': ebitda_margin * 0.85,
                         'wacc': 0.12, 'tg': 0.04},
            }

        results = {}
        for name, s in scenarios.items():
            rev = [rev_cr]
            rev.append(rev[-1] * (1 + s['g1']))
            rev.append(rev[-1] * (1 + s['g2']))
            fcfs = []
            for r in rev:
                eb = r * s['margin']
                dep = r * 0.03
                ebit = eb - dep
                nopat = ebit * (1 - 0.252)
                f = nopat + dep - r * 0.05 - r * 0.08
                fcfs.append(f)
            tv = fcfs[-1] * (1 + s['tg']) / max(s['wacc'] - s['tg'], 0.01)
            pv = sum(f / (1 + s['wacc'])**t for t, f in enumerate(fcfs))
            pv_tv = tv / (1 + s['wacc'])**2
            ev = pv + pv_tv
            eq = ev - net_debt_cr
            iv = eq / max(shares_cr, 0.01)
            upside = (iv / price - 1) if price > 0 else 0
            results[name] = {
                'revenues': rev, 'fcfs': fcfs, 'terminal_value': tv,
                'enterprise_value': ev, 'equity_value': eq,
                'intrinsic_value': iv, 'upside': upside,
                'params': s
            }
        return results

    # --- MONTE CARLO ---
    def monte_carlo_dcf(self, ticker, n_sims=5000):
        info = self.f.data.get(ticker, {})
        revenue = (info.get('revenue', 0) or 0) / 1e7
        ebitda = info.get('ebitda', 0) or 0
        rev_raw = info.get('revenue', 0) or 1
        ebitda_margin = ebitda / rev_raw if rev_raw > 0 else 0.15
        net_debt = ((info.get('total_debt', 0) or 0) - (info.get('total_cash', 0) or 0)) / 1e7
        shares = (info.get('shares_outstanding', 0) or 1) / 1e7
        price = info.get('current_price', 0) or 0
        rev_growth = info.get('revenue_growth', 0) or 0.10

        values = []
        np.random.seed(42)
        for _ in range(n_sims):
            g1 = np.random.normal(rev_growth, abs(rev_growth) * 0.4 + 0.02)
            g2 = np.random.normal(rev_growth * 0.8, abs(rev_growth) * 0.3 + 0.02)
            m = np.random.normal(ebitda_margin, max(ebitda_margin * 0.15, 0.02))
            w = np.random.normal(0.11, 0.015)
            tg = np.random.normal(0.05, 0.01)
            if w <= tg or w <= 0: continue
            revs = [revenue, revenue * (1 + g1)]
            revs.append(revs[-1] * (1 + g2))
            fcfs = []
            for r in revs:
                eb = r * m
                dep = r * 0.03
                ebit = eb - dep
                nopat = ebit * 0.748
                f = nopat + dep - r * 0.05 - r * 0.08
                fcfs.append(f)
            tv = fcfs[-1] * (1 + tg) / (w - tg)
            pv = sum(f / (1 + w)**t for t, f in enumerate(fcfs))
            pv_tv = tv / (1 + w)**2
            eq = pv + pv_tv - net_debt
            iv = eq / max(shares, 0.01)
            if 0 < iv < price * 10:
                values.append(iv)

        if not values:
            return {'mean': 0, 'median': 0, 'p10': 0, 'p25': 0, 'p75': 0, 'p90': 0,
                    'std': 0, 'prob_upside': 0, 'n_valid': 0}
        v = np.array(values)
        return {
            'mean': np.mean(v), 'median': np.median(v),
            'p10': np.percentile(v, 10), 'p25': np.percentile(v, 25),
            'p75': np.percentile(v, 75), 'p90': np.percentile(v, 90),
            'std': np.std(v),
            'prob_upside': (v > price).mean() if price > 0 else 0,
            'n_valid': len(v)
        }

    # --- ESG SCORING ---
    def esg_score(self, ticker):
        info = self.f.data.get(ticker, {})
        sector = info.get('sector', '')
        # Heuristic ESG scoring based on sector & fundamentals
        sector_e = {'Technology': 6, 'Industrials': 5, 'Energy': 4, 'Utilities': 7,
                    'Consumer Cyclical': 6, 'Financial Services': 6, 'Basic Materials': 4,
                    'Healthcare': 7, 'Communication Services': 6, 'Consumer Defensive': 7,
                    'Real Estate': 5}.get(sector, 5)
        roe = info.get('roe', 0) or 0
        de = info.get('debt_to_equity', 0) or 0
        # Governance heuristic: low debt, high ROE = better governance proxy
        g_score = min(9, max(3, 7 + int(roe > 0.15) - int(de > 100)))
        s_score = min(9, max(3, 6 + int(sector in ['Healthcare', 'Technology', 'Consumer Defensive'])))
        fin_score = min(9, max(3, 5 + int(roe > 0.15) + int(roe > 0.25) + int(de < 50)))
        sdg_score = min(9, max(3, sector_e + 1))
        weighted_esg = sector_e * 0.3 + s_score * 0.3 + g_score * 0.4
        composite = fin_score * 0.4 + weighted_esg * 0.35 + sdg_score * 0.25
        if composite >= 7.5: rec = 'STRONG BUY'
        elif composite >= 6.5: rec = 'BUY'
        elif composite >= 5.5: rec = 'HOLD'
        else: rec = 'AVOID'
        return {'E': sector_e, 'S': s_score, 'G': g_score, 'SDG': sdg_score,
                'Financial': fin_score, 'Weighted_ESG': weighted_esg,
                'Composite': composite, 'Recommendation': rec}

    # --- TECHNICAL SIGNALS ---
    def technical_signals(self, ticker):
        if self.f.prices is None or self.f.prices.empty:
            return {}
        if ticker not in self.f.prices.columns:
            return {}
        p = self.f.prices[ticker].dropna()
        if len(p) < 200:
            return {}
        current = p.iloc[-1]
        sma50 = p.rolling(50).mean().iloc[-1]
        sma200 = p.rolling(200).mean().iloc[-1]
        rsi_delta = p.diff()
        gain = rsi_delta.clip(lower=0).rolling(14).mean().iloc[-1]
        loss = (-rsi_delta.clip(upper=0)).rolling(14).mean().iloc[-1]
        rs = gain / max(loss, 1e-10)
        rsi = 100 - 100 / (1 + rs)
        high_52 = p.tail(252).max()
        low_52 = p.tail(252).min()
        pct_from_high = (current / high_52 - 1)
        pct_from_low = (current / low_52 - 1)
        # MACD
        ema12 = p.ewm(span=12).mean().iloc[-1]
        ema26 = p.ewm(span=26).mean().iloc[-1]
        macd = ema12 - ema26
        signal = p.ewm(span=12).mean().ewm(span=9).mean().iloc[-1] - p.ewm(span=26).mean().ewm(span=9).mean().iloc[-1]
        # Momentum score (composite)
        mom_score = 0
        if current > sma50: mom_score += 2
        if current > sma200: mom_score += 2
        if sma50 > sma200: mom_score += 2  # Golden cross
        if rsi < 30: mom_score += 2  # Oversold
        elif rsi > 70: mom_score -= 1  # Overbought
        if macd > 0: mom_score += 1
        mom_score = max(1, min(10, mom_score + 3))  # Scale 1-10
        return {
            'current_price': current, 'sma_50': sma50, 'sma_200': sma200,
            'rsi_14': rsi, 'macd': macd,
            '52w_high': high_52, '52w_low': low_52,
            'pct_from_high': pct_from_high, 'pct_from_low': pct_from_low,
            'momentum_score': mom_score,
            'trend': 'BULLISH' if current > sma200 else 'BEARISH',
            'rsi_signal': 'OVERSOLD' if rsi < 30 else ('OVERBOUGHT' if rsi > 70 else 'NEUTRAL'),
        }


# ============================================================
# EXCEL REPORT GENERATOR
# ============================================================
class ExcelReportGenerator:
    """Generates professional Excel workbook from model results."""

    def __init__(self, fetcher, models):
        self.f = fetcher
        self.m = models
        self.wb = Workbook()

    def generate(self, output_path):
        print(f"\n{'='*60}")
        print(f"  GENERATING EXCEL REPORT")
        print(f"{'='*60}")

        self._create_dashboard()
        self._create_fundamentals()
        self._create_dcf()
        self._create_factor_model()
        self._create_portfolio()
        self._create_garch_risk()
        self._create_esg()
        self._create_technicals()

        # Remove default sheet if still there
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self.wb.save(output_path)
        print(f"\n  âœ… Report saved: {output_path}")
        return output_path

    def _create_dashboard(self):
        ws = self.wb.create_sheet("Dashboard", 0)
        ws.sheet_properties.tabColor = "1F4E79"

        ws['A1'] = "NSE STOCK INVESTMENT ANALYSIS"
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')
        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')} | 1-2 Year Horizon"
        ws['A2'].font = SUBTITLE_FONT

        ws['A4'] = "STOCKS ANALYZED"
        ws['A4'].font = SECTION_FONT

        headers = ["#", "Ticker", "Company", "Sector", "CMP (â‚¹)", "Market Cap (â‚¹ Cr)",
                   "P/E", "ROE", "Revenue Growth", "Momentum"]
        for j, h in enumerate(headers):
            ws.cell(row=5, column=j + 1, value=h)
        style_header_row(ws, 5, len(headers))

        for i, ticker in enumerate(self.f.tickers):
            r = 6 + i
            info = self.f.data.get(ticker, {})
            tech = self.m.technical_signals(ticker)

            vals = [
                i + 1,
                ticker.replace('.NS', ''),
                info.get('name', 'N/A'),
                info.get('sector', 'N/A'),
                info.get('current_price', 0),
                (info.get('market_cap', 0) or 0) / 1e7,
                info.get('pe_ratio', 0) or 0,
                info.get('roe', 0) or 0,
                info.get('revenue_growth', 0) or 0,
                tech.get('momentum_score', 'N/A'),
            ]
            fmts = [NUM_FMT, None, None, None, CUR_FMT, NUM_FMT, DEC_FMT, PCT_FMT, PCT_FMT, None]

            for j, (val, fmt) in enumerate(zip(vals, fmts)):
                cell = style_cell(ws, r, j + 1, font=BLACK_FONT, fmt=fmt)
                cell.value = val
                if j == 0: cell.font = BLACK_BOLD

        # Model Summary
        r_start = 6 + len(self.f.tickers) + 2
        ws.cell(row=r_start, column=1, value="MODELS INCLUDED").font = SECTION_FONT
        models_list = [
            "Fundamentals Overview â€” Key financial metrics from live NSE data",
            "DCF Valuation â€” 3-scenario DCF with Monte Carlo simulation",
            "Factor Model â€” Fama-French 5-Factor expected returns",
            "Portfolio Optimizer â€” Black-Litterman optimal allocation",
            "GARCH Risk â€” Volatility forecast, Kelly sizing, VaR",
            "ESG Scorer â€” Environmental, Social, Governance composite",
            "Technical Analysis â€” Momentum, RSI, MACD, trend signals",
        ]
        for i, m in enumerate(models_list):
            ws.cell(row=r_start + 1 + i, column=1, value=f"  {i+1}. {m}").font = BLACK_FONT

        # Color legend
        r_leg = r_start + len(models_list) + 3
        ws.cell(row=r_leg, column=1, value="COLOR CODING").font = SECTION_FONT
        legends = [
            ("Blue text / Yellow bg", "Editable inputs", BLUE_FONT, INPUT_FILL),
            ("Black text", "Formulas (auto-calculated)", BLACK_FONT, None),
            ("Blue background", "Computed results", BLACK_FONT, RESULT_FILL),
            ("Green background", "Key outputs & recommendations", BLACK_BOLD, GREEN_FILL),
        ]
        for i, (label, desc, fnt, fl) in enumerate(legends):
            c1 = ws.cell(row=r_leg + 1 + i, column=1, value=label)
            c1.font = fnt
            if fl: c1.fill = fl
            ws.cell(row=r_leg + 1 + i, column=2, value=desc).font = BLACK_FONT

        auto_width(ws)
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 22

    def _create_fundamentals(self):
        ws = self.wb.create_sheet("Fundamentals")
        ws.sheet_properties.tabColor = "2E86C1"

        ws['A1'] = "FUNDAMENTAL ANALYSIS â€” LIVE NSE DATA"
        ws['A1'].font = TITLE_FONT

        tickers = self.f.tickers
        headers = ["Metric"] + [t.replace('.NS', '') for t in tickers]
        for j, h in enumerate(headers):
            ws.cell(row=3, column=j + 1, value=h)
        style_header_row(ws, 3, len(headers))

        metrics = [
            ("Company Name", 'name', None),
            ("Sector", 'sector', None),
            ("Industry", 'industry', None),
            ("", None, None),
            ("Current Price (â‚¹)", 'current_price', CUR_FMT),
            ("Market Cap (â‚¹ Cr)", 'market_cap', NUM_FMT),
            ("Enterprise Value (â‚¹ Cr)", 'enterprise_value', NUM_FMT),
            ("52-Week High (â‚¹)", 'fifty_two_week_high', CUR_FMT),
            ("52-Week Low (â‚¹)", 'fifty_two_week_low', CUR_FMT),
            ("", None, None),
            ("P/E Ratio (TTM)", 'pe_ratio', DEC_FMT),
            ("Forward P/E", 'forward_pe', DEC_FMT),
            ("P/B Ratio", 'pb_ratio', DEC_FMT),
            ("EV/EBITDA", 'ev_ebitda', DEC_FMT),
            ("Dividend Yield", 'dividend_yield', PCT_FMT),
            ("", None, None),
            ("Revenue (â‚¹ Cr)", 'revenue', NUM_FMT),
            ("EBITDA (â‚¹ Cr)", 'ebitda', NUM_FMT),
            ("Net Income (â‚¹ Cr)", 'net_income', NUM_FMT),
            ("Free Cash Flow (â‚¹ Cr)", 'free_cashflow', NUM_FMT),
            ("", None, None),
            ("Revenue Growth", 'revenue_growth', PCT_FMT),
            ("Earnings Growth", 'earnings_growth', PCT_FMT),
            ("Gross Margin", 'gross_margin', PCT_FMT),
            ("Operating Margin", 'operating_margin', PCT_FMT),
            ("Profit Margin", 'profit_margin', PCT_FMT),
            ("ROE", 'roe', PCT_FMT),
            ("", None, None),
            ("Beta", 'beta', DEC_FMT),
            ("Debt/Equity", 'debt_to_equity', DEC_FMT),
            ("Total Debt (â‚¹ Cr)", 'total_debt', NUM_FMT),
            ("Total Cash (â‚¹ Cr)", 'total_cash', NUM_FMT),
            ("Book Value (â‚¹)", 'book_value', CUR_FMT),
        ]

        for i, (label, key, fmt) in enumerate(metrics):
            r = 4 + i
            ws.cell(row=r, column=1, value=label).font = BLACK_BOLD if label else BLACK_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER

            if key:
                for j, ticker in enumerate(tickers):
                    info = self.f.data.get(ticker, {})
                    val = info.get(key, 0)
                    # Convert large numbers to Crores
                    if key in ['market_cap', 'enterprise_value', 'revenue', 'ebitda',
                               'net_income', 'free_cashflow', 'total_debt', 'total_cash']:
                        val = (val or 0) / 1e7
                    cell = style_cell(ws, r, j + 2, font=BLACK_FONT, fmt=fmt)
                    cell.value = val if val is not None else ''

        auto_width(ws)
        ws.column_dimensions['A'].width = 26

    def _create_dcf(self):
        ws = self.wb.create_sheet("DCF Valuation")
        ws.sheet_properties.tabColor = "E74C3C"

        ws['A1'] = "DCF VALUATION â€” ALL STOCKS"
        ws['A1'].font = TITLE_FONT

        row = 3
        for ticker in self.f.tickers:
            info = self.f.data.get(ticker, {})
            name = info.get('name', ticker.replace('.NS', ''))
            price = info.get('current_price', 0)

            ws.cell(row=row, column=1, value=f"â–¶ {name} ({ticker.replace('.NS', '')})").font = SECTION_FONT
            row += 1

            dcf = self.m.dcf_valuation(ticker)
            mc = self.m.monte_carlo_dcf(ticker)

            # Scenario table
            headers = ["", "Base", "Bull", "Bear"]
            for j, h in enumerate(headers):
                ws.cell(row=row, column=j + 1, value=h)
            style_header_row(ws, row, 4)
            row += 1

            params_display = [
                ("Revenue Growth Y1", 'g1', PCT_FMT),
                ("Revenue Growth Y2", 'g2', PCT_FMT),
                ("EBITDA Margin", 'margin', PCT_FMT),
                ("WACC", 'wacc', PCT_FMT),
                ("Terminal Growth", 'tg', PCT_FMT),
                ("", None, None),
                ("Intrinsic Value (â‚¹)", None, CUR_FMT),
                ("Current Price (â‚¹)", None, CUR_FMT),
                ("Upside / Downside", None, PCT_FMT),
            ]

            for label, key, fmt in params_display:
                ws.cell(row=row, column=1, value=label).font = BLACK_BOLD
                ws.cell(row=row, column=1).border = THIN_BORDER

                for j, scenario in enumerate(['Base', 'Bull', 'Bear']):
                    cell = style_cell(ws, row, j + 2, fmt=fmt)
                    if key and scenario in dcf:
                        cell.value = dcf[scenario]['params'].get(key, 0)
                        cell.font = BLUE_FONT
                        cell.fill = INPUT_FILL
                    elif label == "Intrinsic Value (â‚¹)" and scenario in dcf:
                        cell.value = dcf[scenario]['intrinsic_value']
                        cell.font = BLACK_BOLD
                        cell.fill = GREEN_FILL
                    elif label == "Current Price (â‚¹)":
                        cell.value = price
                        cell.font = BLACK_FONT
                    elif label == "Upside / Downside" and scenario in dcf:
                        cell.value = dcf[scenario]['upside']
                        cell.font = BLACK_BOLD
                        cell.fill = RESULT_FILL
                    elif label == "":
                        pass
                row += 1

            # Monte Carlo results
            ws.cell(row=row, column=1, value="Monte Carlo (5,000 sims)").font = SECTION_FONT
            row += 1
            mc_items = [
                ("Mean Value (â‚¹)", mc.get('mean', 0), CUR_FMT),
                ("Median Value (â‚¹)", mc.get('median', 0), CUR_FMT),
                ("10th Percentile (â‚¹)", mc.get('p10', 0), CUR_FMT),
                ("90th Percentile (â‚¹)", mc.get('p90', 0), CUR_FMT),
                ("P(Upside > 0%)", mc.get('prob_upside', 0), PCT_FMT),
            ]
            for label, val, fmt in mc_items:
                ws.cell(row=row, column=1, value=label).font = BLACK_FONT
                ws.cell(row=row, column=1).border = THIN_BORDER
                cell = style_cell(ws, row, 2, font=BLACK_BOLD, fmt=fmt, fill=RESULT_FILL)
                cell.value = val
                row += 1

            row += 2

        auto_width(ws)
        ws.column_dimensions['A'].width = 28

    def _create_factor_model(self):
        ws = self.wb.create_sheet("Factor Model")
        ws.sheet_properties.tabColor = "28B463"

        ws['A1'] = "FAMA-FRENCH 5-FACTOR MODEL"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "E(Ri) = Rf + Î²m(Rm-Rf) + Î²sÂ·SMB + Î²vÂ·HML + Î²rÂ·RMW + Î²cÂ·CMA"
        ws['A2'].font = SUBTITLE_FONT

        # Market parameters
        ws['A4'] = "MARKET PARAMETERS"
        ws['A4'].font = SECTION_FONT
        params = [
            ("Risk-Free Rate (Rf)", 0.071, "India 10Y Govt Bond"),
            ("Market Risk Premium", 0.059, "NIFTY 50 avg - Rf"),
            ("SMB Premium", 0.025, "Size"),
            ("HML Premium", 0.035, "Value"),
            ("RMW Premium", 0.030, "Profitability"),
            ("CMA Premium", 0.020, "Investment"),
        ]
        for j, h in enumerate(["Parameter", "Value", "Description"]):
            ws.cell(row=5, column=j + 1, value=h)
        style_header_row(ws, 5, 3)
        for i, (p, v, d) in enumerate(params):
            r = 6 + i
            ws.cell(row=r, column=1, value=p).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            cell = style_cell(ws, r, 2, font=BLUE_FONT, fill=INPUT_FILL, fmt=PCT_FMT)
            cell.value = v
            ws.cell(row=r, column=3, value=d).font = SMALL_GRAY
            ws.cell(row=r, column=3).border = THIN_BORDER

        # Factor loadings table
        r_start = 14
        ws.cell(row=r_start, column=1, value="STOCK FACTOR LOADINGS & EXPECTED RETURNS").font = SECTION_FONT
        headers = ["Stock", "Î² Market", "Î² Size", "Î² Value", "Î² Profit", "Î² Invest", "E(Return)", "Excess Return"]
        for j, h in enumerate(headers):
            ws.cell(row=r_start + 1, column=j + 1, value=h)
        style_header_row(ws, r_start + 1, len(headers))

        for i, ticker in enumerate(self.f.tickers):
            r = r_start + 2 + i
            betas = self.m.estimate_factor_betas(ticker)
            er = self.m.factor_expected_return(betas)

            ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER

            for j, key in enumerate(['market', 'smb', 'hml', 'rmw', 'cma']):
                cell = style_cell(ws, r, j + 2, font=BLUE_FONT, fill=INPUT_FILL, fmt=DEC_FMT)
                cell.value = betas[key]

            cell = style_cell(ws, r, 7, font=BLACK_BOLD, fill=GREEN_FILL, fmt=PCT_FMT)
            cell.value = er
            cell = style_cell(ws, r, 8, font=BLACK_FONT, fill=RESULT_FILL, fmt=PCT_FMT)
            cell.value = er - 0.071

        auto_width(ws)
        ws.column_dimensions['A'].width = 20

    def _create_portfolio(self):
        ws = self.wb.create_sheet("Portfolio Optimizer")
        ws.sheet_properties.tabColor = "8E44AD"

        ws['A1'] = "BLACK-LITTERMAN PORTFOLIO OPTIMIZATION"
        ws['A1'].font = TITLE_FONT

        # Generate views from factor model
        views = []
        for ticker in self.f.tickers:
            betas = self.m.estimate_factor_betas(ticker)
            er = self.m.factor_expected_return(betas)
            views.append({'ticker': ticker, 'return': er, 'confidence': 0.6})

        bl = self.m.black_litterman(views)

        if not bl:
            ws['A3'] = "Insufficient data for portfolio optimization"
            ws['A3'].font = Font(name='Arial', color='FF0000', size=12, bold=True)
            return

        # Parameters
        ws['A3'] = "PORTFOLIO PARAMETERS"
        ws['A3'].font = SECTION_FONT
        for i, (p, v) in enumerate([("Risk Aversion (Î´)", 2.5), ("Tau", 0.05),
                                     ("Portfolio Value (â‚¹ Lakhs)", 50), ("Max Position", "40%"), ("Min Position", "2%")]):
            r = 4 + i
            ws.cell(row=r, column=1, value=p).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            cell = style_cell(ws, r, 2, font=BLUE_FONT, fill=INPUT_FILL)
            cell.value = v

        # Allocation table
        r_start = 10
        ws.cell(row=r_start, column=1, value="OPTIMAL ALLOCATION").font = SECTION_FONT
        headers = ["Stock", "Mkt Cap Weight", "Equilibrium E(R)", "BL E(R)", "Optimal Weight", "Allocation (â‚¹ L)"]
        for j, h in enumerate(headers):
            ws.cell(row=r_start + 1, column=j + 1, value=h)
        style_header_row(ws, r_start + 1, len(headers))

        tickers = bl['tickers']
        for i, ticker in enumerate(tickers):
            r = r_start + 2 + i
            ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER

            cell = style_cell(ws, r, 2, font=BLACK_FONT, fmt=PCT_FMT)
            cell.value = bl['mcap_weights'].get(ticker, 0)
            cell = style_cell(ws, r, 3, font=BLACK_FONT, fmt=PCT_FMT)
            cell.value = bl['equilibrium'].get(ticker, 0)
            cell = style_cell(ws, r, 4, font=BLACK_FONT, fmt=PCT_FMT, fill=RESULT_FILL)
            cell.value = bl['bl_returns'].get(ticker, 0)
            cell = style_cell(ws, r, 5, font=BLACK_BOLD, fmt=PCT_FMT, fill=GREEN_FILL)
            cell.value = bl['optimal_weights'].get(ticker, 0)
            cell = style_cell(ws, r, 6, font=BLACK_BOLD, fmt=DEC_FMT, fill=GREEN_FILL)
            cell.value = bl['optimal_weights'].get(ticker, 0) * 50

        # Summary
        r_sum = r_start + 2 + len(tickers) + 2
        ws.cell(row=r_sum, column=1, value="PORTFOLIO SUMMARY").font = SECTION_FONT
        summary = [
            ("Expected Return", bl['port_return'], PCT_FMT),
            ("Portfolio Volatility", bl['port_vol'], PCT_FMT),
            ("Sharpe Ratio", bl['sharpe'], DEC_FMT),
            ("Est. Max Drawdown", bl['port_vol'] * -2.5, PCT_FMT),
        ]
        for i, (label, val, fmt) in enumerate(summary):
            r = r_sum + 1 + i
            ws.cell(row=r, column=1, value=label).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            cell = style_cell(ws, r, 2, font=BLACK_BOLD, fmt=fmt, fill=GREEN_FILL)
            cell.value = val

        auto_width(ws)
        ws.column_dimensions['A'].width = 24

    def _create_garch_risk(self):
        ws = self.wb.create_sheet("GARCH Risk")
        ws.sheet_properties.tabColor = "E74C3C"

        ws['A1'] = "GARCH(1,1) VOLATILITY & POSITION SIZING"
        ws['A1'].font = TITLE_FONT
        ws['A2'] = "ÏƒÂ²(t) = Ï‰ + Î±Â·ÎµÂ²(t-1) + Î²Â·ÏƒÂ²(t-1)"
        ws['A2'].font = SUBTITLE_FONT

        headers = ["Stock", "Ï‰", "Î± (ARCH)", "Î² (GARCH)", "Persistence",
                   "LR Vol (ann)", "Current Vol", "Kelly %", "Half Kelly",
                   "VaR 95%", "VaR 99%", "Max Position (â‚¹L)"]
        ws.cell(row=4, column=1, value="GARCH PARAMETERS & POSITION SIZING").font = SECTION_FONT
        for j, h in enumerate(headers):
            ws.cell(row=5, column=j + 1, value=h)
        style_header_row(ws, 5, len(headers))

        for i, ticker in enumerate(self.f.tickers):
            r = 6 + i
            garch = self.m.fit_garch(ticker)
            betas = self.m.estimate_factor_betas(ticker)
            er = self.m.factor_expected_return(betas)
            vols = self.f.get_volatilities()
            vol = vols.get(ticker, 0.25)

            persistence = garch['alpha'] + garch['beta']
            lr_var = garch['omega'] / max(1 - persistence, 0.001)
            lr_vol = np.sqrt(252 * lr_var)

            kelly = (er - 0.071) / max(vol**2, 0.001)
            half_kelly = kelly / 2
            var95 = -(er - 1.645 * vol) * 50
            var99 = -(er - 2.326 * vol) * 50
            max_pos = min(max(half_kelly, 0.02), 0.40) * 50

            ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER

            vals = [garch['omega'], garch['alpha'], garch['beta'], persistence,
                    lr_vol, vol, kelly, half_kelly, var95, var99, max_pos]
            fmts_row = ['0.00000', DEC_FMT, DEC_FMT, DEC_FMT,
                        PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT,
                        DEC_FMT, DEC_FMT, DEC_FMT]

            for j, (val, fmt) in enumerate(zip(vals, fmts_row)):
                cell = style_cell(ws, r, j + 2, font=BLACK_FONT, fmt=fmt)
                cell.value = val
                if j >= 6:  # Results
                    cell.fill = RESULT_FILL
                if j == 10:  # Max position
                    cell.fill = GREEN_FILL
                    cell.font = BLACK_BOLD

        # Volatility forecast section
        r_fc = 6 + len(self.f.tickers) + 2
        ws.cell(row=r_fc, column=1, value="VOLATILITY TERM STRUCTURE FORECAST").font = SECTION_FONT

        fc_headers = ["Stock", "5-day", "21-day (1M)", "63-day (3M)", "126-day (6M)", "252-day (1Y)"]
        for j, h in enumerate(fc_headers):
            ws.cell(row=r_fc + 1, column=j + 1, value=h)
        style_header_row(ws, r_fc + 1, len(fc_headers))

        for i, ticker in enumerate(self.f.tickers):
            r = r_fc + 2 + i
            garch = self.m.fit_garch(ticker)
            vols_dict = self.f.get_volatilities()
            current_daily_var = (vols_dict.get(ticker, 0.25) / np.sqrt(252))**2
            forecasts = self.m.garch_forecast(garch, current_daily_var)

            ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER

            for j, h in enumerate([4, 20, 62, 125, 251]):
                ann_vol = np.sqrt(252 * forecasts[h]) if h < len(forecasts) else 0
                cell = style_cell(ws, r, j + 2, font=BLACK_FONT, fmt=PCT_FMT, fill=RESULT_FILL)
                cell.value = ann_vol

        auto_width(ws)
        ws.column_dimensions['A'].width = 18

    def _create_esg(self):
        ws = self.wb.create_sheet("ESG Scorer")
        ws.sheet_properties.tabColor = "27AE60"

        ws['A1'] = "ESG-INTEGRATED INVESTMENT SCORING"
        ws['A1'].font = TITLE_FONT

        ws['A3'] = "PILLAR WEIGHTS"
        ws['A3'].font = SECTION_FONT
        for i, (p, w) in enumerate([("Environmental (E)", 0.30), ("Social (S)", 0.30), ("Governance (G)", 0.40)]):
            r = 4 + i
            ws.cell(row=r, column=1, value=p).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            cell = style_cell(ws, r, 2, font=BLUE_FONT, fill=INPUT_FILL, fmt=PCT_FMT)
            cell.value = w

        ws['A8'] = "COMPOSITE WEIGHTS: Financial 40% | ESG 35% | SDG 25%"
        ws['A8'].font = SMALL_GRAY

        headers = ["Stock", "E Score", "S Score", "G Score", "Weighted ESG",
                   "SDG Align", "Financial", "Composite", "Recommendation"]
        ws.cell(row=10, column=1, value="STOCK ESG SCORES (1-10 scale)").font = SECTION_FONT
        for j, h in enumerate(headers):
            ws.cell(row=11, column=j + 1, value=h)
        style_header_row(ws, 11, len(headers))

        for i, ticker in enumerate(self.f.tickers):
            r = 12 + i
            esg = self.m.esg_score(ticker)

            ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER

            vals = [esg['E'], esg['S'], esg['G'], esg['Weighted_ESG'],
                    esg['SDG'], esg['Financial'], esg['Composite'], esg['Recommendation']]
            for j, val in enumerate(vals):
                if j < 3:
                    cell = style_cell(ws, r, j + 2, font=BLUE_FONT, fill=INPUT_FILL)
                elif j < 6:
                    cell = style_cell(ws, r, j + 2, font=BLACK_FONT, fill=RESULT_FILL, fmt=DEC_FMT)
                elif j == 6:
                    cell = style_cell(ws, r, j + 2, font=BLACK_BOLD, fill=GREEN_FILL, fmt=DEC_FMT)
                else:
                    cell = style_cell(ws, r, j + 2, font=BLACK_BOLD, fill=GREEN_FILL, align=CENTER)
                    # Color code recommendation
                    if val == 'STRONG BUY':
                        cell.fill = GREEN_FILL
                    elif val == 'BUY':
                        cell.fill = RESULT_FILL
                    elif val == 'HOLD':
                        cell.fill = INPUT_FILL
                    else:
                        cell.fill = RED_FILL
                cell.value = val

        auto_width(ws)
        ws.column_dimensions['A'].width = 18

    def _create_technicals(self):
        ws = self.wb.create_sheet("Technicals")
        ws.sheet_properties.tabColor = "D35400"

        ws['A1'] = "TECHNICAL ANALYSIS â€” MOMENTUM & TREND SIGNALS"
        ws['A1'].font = TITLE_FONT

        headers = ["Stock", "CMP (â‚¹)", "SMA 50", "SMA 200", "RSI (14)",
                   "MACD", "52W High", "52W Low", "% from High", "% from Low",
                   "Trend", "RSI Signal", "Momentum (1-10)"]
        for j, h in enumerate(headers):
            ws.cell(row=3, column=j + 1, value=h)
        style_header_row(ws, 3, len(headers))

        for i, ticker in enumerate(self.f.tickers):
            r = 4 + i
            tech = self.m.technical_signals(ticker)
            if not tech:
                ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
                ws.cell(row=r, column=2, value="Insufficient data").font = SMALL_GRAY
                continue

            ws.cell(row=r, column=1, value=ticker.replace('.NS', '')).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER

            vals = [
                (tech.get('current_price', 0), CUR_FMT),
                (tech.get('sma_50', 0), CUR_FMT),
                (tech.get('sma_200', 0), CUR_FMT),
                (tech.get('rsi_14', 0), DEC_FMT),
                (tech.get('macd', 0), DEC_FMT),
                (tech.get('52w_high', 0), CUR_FMT),
                (tech.get('52w_low', 0), CUR_FMT),
                (tech.get('pct_from_high', 0), PCT_FMT),
                (tech.get('pct_from_low', 0), PCT_FMT),
                (tech.get('trend', ''), None),
                (tech.get('rsi_signal', ''), None),
                (tech.get('momentum_score', 0), None),
            ]

            for j, (val, fmt) in enumerate(vals):
                cell = style_cell(ws, r, j + 2, font=BLACK_FONT, fmt=fmt)
                cell.value = val

                # Color code signals
                if j == 9:  # Trend
                    cell.fill = GREEN_FILL if val == 'BULLISH' else RED_FILL
                    cell.font = BLACK_BOLD
                elif j == 10:  # RSI
                    if val == 'OVERSOLD': cell.fill = GREEN_FILL
                    elif val == 'OVERBOUGHT': cell.fill = RED_FILL
                    cell.font = BLACK_BOLD
                elif j == 11:  # Momentum
                    if val >= 7: cell.fill = GREEN_FILL
                    elif val <= 4: cell.fill = RED_FILL
                    else: cell.fill = INPUT_FILL
                    cell.font = BLACK_BOLD

        # Signal interpretation guide
        r_guide = 4 + len(self.f.tickers) + 2
        ws.cell(row=r_guide, column=1, value="SIGNAL INTERPRETATION").font = SECTION_FONT
        guide = [
            ("Momentum 8-10", "Strong uptrend â€” consider entry on pullbacks"),
            ("Momentum 5-7", "Neutral â€” wait for clearer direction"),
            ("Momentum 1-4", "Downtrend â€” avoid or wait for reversal signals"),
            ("RSI < 30", "Oversold â€” potential mean reversion buy opportunity"),
            ("RSI > 70", "Overbought â€” consider partial profit taking"),
            ("Price > SMA200", "Long-term uptrend intact"),
            ("SMA50 > SMA200", "Golden Cross â€” bullish medium-term signal"),
        ]
        for i, (sig, desc) in enumerate(guide):
            r = r_guide + 1 + i
            ws.cell(row=r, column=1, value=sig).font = BLACK_BOLD
            ws.cell(row=r, column=1).border = THIN_BORDER
            ws.cell(row=r, column=2, value=desc).font = BLACK_FONT
            ws.cell(row=r, column=2).border = THIN_BORDER

        auto_width(ws)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50


# ============================================================
# MAIN EXECUTION
# ============================================================
def load_tickers_from_file(file_path: str):
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Ticker file not found: {path}")

    if path.suffix.lower() == '.csv':
        df = pd.read_csv(path)
        if df.empty:
            return []
        cols = {c.upper().strip(): c for c in df.columns}
        source_col = cols.get('SYMBOL') or df.columns[0]
        vals = df[source_col].dropna().astype(str).tolist()
        return [v.strip().upper().replace('.NS', '') for v in vals if v.strip()]

    raw = path.read_text(encoding='utf-8')
    tokens = raw.replace(',', ' ').split()
    cleaned = []
    for token in tokens:
        t = token.strip()
        if not t or t.startswith('#'):
            continue
        cleaned.append(t.upper().replace('.NS', ''))
    return cleaned


def select_mode_interactive():
    print("\nSelect search approach:")
    print("  1) Selected Stock Search (single ticker)")
    print("  2) Multiple Stock Search (space/comma separated)")
    print("  3) Load Multiple Stocks From File")
    choice = input("\n  Enter 1, 2, or 3: ").strip()

    if choice == '1':
        ticker = input("\n  Enter ticker (e.g., RELIANCE): ").strip().upper()
        return [ticker] if ticker else []

    if choice == '3':
        file_path = input("\n  Enter file path (.txt or .csv): ").strip()
        try:
            return load_tickers_from_file(file_path)
        except Exception as e:
            print(f"❌ Could not load ticker file: {e}")
            return []

    print("Enter NSE stock tickers (comma or space separated):")
    print("Examples: HAL, BEL, RELIANCE, TATAPOWER, INFY")
    raw = input("\n  > ").strip()
    return [t.strip().upper() for t in raw.replace(',', ' ').split() if t.strip()]


def parse_tickers_from_args():
    parser = argparse.ArgumentParser(
        description='NSE Stock Investment Analyzer',
        add_help=True,
    )
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument('--single', type=str, help='Analyze one ticker (selected stock search)')
    mode_group.add_argument('--multi', nargs='+', help='Analyze multiple tickers')
    mode_group.add_argument('--ticker-file', type=str, help='Load tickers from .txt/.csv file')
    parser.add_argument('--screener', action='store_true', help='Run multi-model stock screener and rank top stocks')
    parser.add_argument('--top-n', type=int, default=10, help='Top N stocks to display in screener mode (default: 10)')
    parser.add_argument('--interactive', action='store_true', help='Interactive mode with search approach selection')
    parser.add_argument('tickers', nargs='*', help='Legacy direct ticker arguments (treated as multiple search)')

    args = parser.parse_args()

    if args.interactive:
        return select_mode_interactive(), args
    if args.single:
        return [args.single.strip().upper().replace('.NS', '')], args
    if args.multi:
        return [t.strip().upper().replace('.NS', '') for t in args.multi if t.strip()], args
    if args.ticker_file:
        return load_tickers_from_file(args.ticker_file), args
    if args.tickers:
        return [t.strip().upper().replace('.NS', '') for t in args.tickers if t.strip()], args

    return select_mode_interactive(), args


def run_analysis(tickers):
    tickers = [t.strip().upper().replace('.NS', '') for t in tickers if t and t.strip()]
    tickers = list(dict.fromkeys(tickers))

    if not tickers:
        print("❌ No tickers provided. Exiting.")
        sys.exit(1)

    print(f"\n  Analyzing: {', '.join(tickers)}")

    fetcher = NSEDataFetcher(tickers)
    fetcher.fetch_all()

    models = QuantModels(fetcher)

    print(f"\n{'='*60}")
    print("  QUICK SUMMARY")
    print(f"{'='*60}")
    for ticker in fetcher.tickers:
        info = fetcher.data.get(ticker, {})
        name = info.get('name', ticker)
        price = info.get('current_price', 0)
        pe = info.get('pe_ratio', 0)
        roe = info.get('roe', 0)
        betas = models.estimate_factor_betas(ticker)
        er = models.factor_expected_return(betas)
        esg = models.esg_score(ticker)
        tech = models.technical_signals(ticker)

        print(f"\n  {name} ({ticker.replace('.NS', '')})")
        print(f"    CMP: ₹{price:,.0f} | P/E: {pe:.1f} | ROE: {(roe or 0)*100:.1f}%")
        print(f"    Factor E(R): {er:.1%} | Beta: {betas['market']:.2f}")
        print(f"    ESG: {esg['Composite']:.1f}/10 → {esg['Recommendation']}")
        if tech:
            print(f"    Momentum: {tech.get('momentum_score', 'N/A')}/10 | Trend: {tech.get('trend', 'N/A')}")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    ticker_str = '_'.join(t.replace('.NS', '') for t in fetcher.tickers[:5])
    if len(fetcher.tickers) > 5:
        ticker_str += f'_+{len(fetcher.tickers)-5}more'
    filename = f'NSE_Analysis_{ticker_str}_{timestamp}.xlsx'
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = str(REPORTS_DIR / filename)

    generator = ExcelReportGenerator(fetcher, models)
    generator.generate(output_path)
    return output_path, filename


def _clamp(v, lo, hi):
    return max(lo, min(hi, v))


def _scale_01(v, lo, hi):
    if hi <= lo:
        return 0.0
    return _clamp((v - lo) / (hi - lo), 0.0, 1.0)


def run_screener(tickers, top_n=10):
    tickers = [t.strip().upper().replace('.NS', '') for t in tickers if t and t.strip()]
    tickers = list(dict.fromkeys(tickers))
    top_n = max(1, int(top_n))

    if not tickers:
        print("❌ No tickers provided for screener.")
        sys.exit(1)

    print(f"\n  Running Screener On: {len(tickers)} ticker(s)")
    fetcher = NSEDataFetcher(tickers)
    fetcher.fetch_all()
    models = QuantModels(fetcher)

    vols = fetcher.get_volatilities()
    rows = []
    for ticker in fetcher.tickers:
        info = fetcher.data.get(ticker, {})
        dcf = models.dcf_valuation(ticker)
        dcf_upside = dcf.get('Base', {}).get('upside', 0) if isinstance(dcf, dict) else 0
        betas = models.estimate_factor_betas(ticker)
        factor_er = models.factor_expected_return(betas)
        esg = models.esg_score(ticker)
        tech = models.technical_signals(ticker)

        roe = info.get('roe', 0) or 0
        rev_growth = info.get('revenue_growth', 0) or 0
        vol = vols.get(ticker, 0.35)
        momentum = (tech.get('momentum_score', 5) if isinstance(tech, dict) else 5) or 5

        dcf_score = _scale_01(dcf_upside, -0.50, 1.00)
        factor_score = _scale_01(factor_er, 0.05, 0.35)
        esg_score = _scale_01(esg.get('Composite', 5), 0.0, 10.0)
        tech_score = _scale_01(momentum, 1.0, 10.0)
        risk_score = 1.0 - _scale_01(vol, 0.12, 0.60)
        quality_score = 0.6 * _scale_01(roe, 0.0, 0.30) + 0.4 * _scale_01(rev_growth, -0.10, 0.30)

        total_score = (
            0.20 * dcf_score +
            0.20 * factor_score +
            0.15 * esg_score +
            0.15 * tech_score +
            0.15 * risk_score +
            0.15 * quality_score
        ) * 100.0

        rows.append({
            'Ticker': ticker.replace('.NS', ''),
            'Company': info.get('name', ticker.replace('.NS', '')),
            'Sector': info.get('sector', 'N/A'),
            'Score_Total': round(total_score, 2),
            'Score_DCF': round(dcf_score * 100, 2),
            'Score_Factor': round(factor_score * 100, 2),
            'Score_ESG': round(esg_score * 100, 2),
            'Score_Technical': round(tech_score * 100, 2),
            'Score_Risk': round(risk_score * 100, 2),
            'Score_Quality': round(quality_score * 100, 2),
            'DCF_Upside_%': round((dcf_upside or 0) * 100, 2),
            'Factor_ER_%': round((factor_er or 0) * 100, 2),
            'ESG_Composite': round(esg.get('Composite', 0), 2),
            'Momentum': momentum,
            'Volatility_%': round((vol or 0) * 100, 2),
        })

    if not rows:
        print("❌ Screener could not compute scores.")
        sys.exit(1)

    df = pd.DataFrame(rows).sort_values('Score_Total', ascending=False)
    top_df = df.head(top_n).copy()
    top_df.insert(0, 'Rank', range(1, len(top_df) + 1))

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    out_name = f'NSE_Screener_Top{top_n}_{timestamp}.csv'
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = REPORTS_DIR / out_name
    top_df.to_csv(out_path, index=False)

    print(f"\n{'='*80}")
    print(f"  TOP {top_n} STOCKS — MULTI-MODEL SCREENER")
    print(f"{'='*80}")
    print(top_df[['Rank', 'Ticker', 'Company', 'Score_Total', 'Score_DCF', 'Score_Factor',
                  'Score_ESG', 'Score_Technical', 'Score_Risk', 'Score_Quality']].to_string(index=False))
    print(f"\n  ✅ Screener output saved: {out_path}")
    return str(out_path), out_name


def main():
    print("""
+-----------------------------------------------------------+
|        NSE STOCK INVESTMENT ANALYZER v2.0                |
|  Mathematical & ESG-Integrated Financial Models          |
+-----------------------------------------------------------+
    """)

    try:
        tickers, args = parse_tickers_from_args()
    except Exception as e:
        print(f"❌ Argument parsing error: {e}")
        sys.exit(1)

    if args.screener:
        return run_screener(tickers, top_n=args.top_n)
    return run_analysis(tickers)


if __name__ == '__main__':
    output_path, filename = main()
    print(f"\n{'='*60}")
    print("  ANALYSIS COMPLETE")
    print(f"  File: {filename}")
    print(f"{'='*60}\n")
