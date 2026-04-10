import csv
import io
from typing import Any, List, Dict, Optional
from datetime import datetime

def export_csv(data: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> str:
    """Export list of dicts to CSV string."""
    if not data:
        return ""
    columns = columns or list(data[0].keys())
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()

def export_excel(sheets: Dict[str, List[Dict[str, Any]]]) -> bytes:
    """Export multiple sheets to Excel bytes.
    sheets: {"Sheet Name": [rows]}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31 char limit
        if not rows:
            continue

        headers = list(rows[0].keys())

        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                # Handle nested dicts/lists by converting to string
                if isinstance(value, (dict, list)):
                    import json
                    value = json.dumps(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Format numbers
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
                elif isinstance(value, datetime):
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'

        # Auto-width columns
        for col_cells in ws.columns:
            max_length = 0
            column = col_cells[0].column_letter # Get the column name
            for cell in col_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
